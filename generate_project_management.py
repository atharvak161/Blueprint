#!/usr/bin/env python3
"""
Project Management Tool Generator
Creates a comprehensive Excel workbook (.xlsm compatible) for project management
with 12+ collapseable phases, UK holiday-aware scheduling, VBA export macros,
and detailed README instructions.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from datetime import date, timedelta
import os
import zipfile
import shutil

# ─── COLOUR PALETTE ──────────────────────────────────────────────────────────
C_NAVY       = "1F4E79"   # Phase header BG
C_NAVY_MID   = "2E75B6"   # Accent / header bar
C_LIGHT_BLUE = "BDD7EE"   # Alt row / subtask BG
C_DARK_GREY  = "404040"   # Dark text
C_WHITE      = "FFFFFF"
C_YELLOW_HDR = "FFF2CC"   # Summary highlight
C_GREEN      = "00B050"
C_AMBER      = "FFB900"
C_RED        = "FF0000"
C_ORANGE     = "FF6600"
C_GREY_NS    = "A5A5A5"   # Not Started
C_PURPLE     = "7030A0"   # Blocked
C_LIGHT_GRN  = "E2EFDA"   # Completed row tint
C_LIGHT_RED  = "FCE4D6"   # Overdue row tint
C_LIGHT_AMBE = "FFF2CC"   # In Progress row tint
C_SUBTASK_BG = "F5F5F5"
C_PHASE_TXT  = "FFFFFF"
C_SUMM_BG    = "1F4E79"
C_SUMM_ACCNT = "2E75B6"

# ─── STYLE HELPERS ───────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(bold=False, size=11, color="000000", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border_all():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def border_medium_bottom():
    med = Side(style="medium", color="1F4E79")
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=med)

def border_thick():
    med = Side(style="medium", color="1F4E79")
    return Border(left=med, right=med, top=med, bottom=med)

# ─── UK HOLIDAYS (2025 & 2026) ────────────────────────────────────────────────
UK_HOLIDAYS = [
    # 2025
    ("New Year's Day", date(2025, 1, 1), date(2025, 1, 1), ""),
    ("Good Friday", date(2025, 4, 18), date(2025, 4, 18), ""),
    ("Easter Monday", date(2025, 4, 21), date(2025, 4, 21), ""),
    ("Early May Bank Holiday", date(2025, 5, 5), date(2025, 5, 5), ""),
    ("Spring Bank Holiday", date(2025, 5, 26), date(2025, 5, 26), ""),
    ("Summer Bank Holiday", date(2025, 8, 25), date(2025, 8, 25), ""),
    ("Christmas Day", date(2025, 12, 25), date(2025, 12, 25), ""),
    ("Boxing Day", date(2025, 12, 26), date(2025, 12, 26), ""),
    # 2026
    ("New Year's Day", date(2026, 1, 1), date(2026, 1, 1), ""),
    ("Good Friday", date(2026, 4, 3), date(2026, 4, 3), ""),
    ("Easter Monday", date(2026, 4, 6), date(2026, 4, 6), ""),
    ("Early May Bank Holiday", date(2026, 5, 4), date(2026, 5, 4), ""),
    ("Spring Bank Holiday", date(2026, 5, 25), date(2026, 5, 25), ""),
    ("Summer Bank Holiday", date(2026, 8, 31), date(2026, 8, 31), ""),
    ("Christmas Day", date(2026, 12, 25), date(2026, 12, 25), ""),
    ("Boxing Day", date(2026, 12, 26), date(2026, 12, 26), ""),
    # 2027 (partial)
    ("New Year's Day", date(2027, 1, 1), date(2027, 1, 1), ""),
    ("Good Friday", date(2027, 3, 26), date(2027, 3, 26), ""),
    ("Easter Monday", date(2027, 3, 29), date(2027, 3, 29), ""),
    ("Early May Bank Holiday", date(2027, 5, 3), date(2027, 5, 3), ""),
    ("Spring Bank Holiday", date(2027, 5, 31), date(2027, 5, 31), ""),
    ("Summer Bank Holiday", date(2027, 8, 30), date(2027, 8, 30), ""),
    ("Christmas Day", date(2027, 12, 27), date(2027, 12, 27), ""),
    ("Boxing Day", date(2027, 12, 28), date(2027, 12, 28), ""),
]

HOLIDAY_DATES = {h[1] for h in UK_HOLIDAYS}

# ─── RESOURCES ───────────────────────────────────────────────────────────────
RESOURCES = [
    ("James Thornton",    "PMO",          "Project Manager"),
    ("Sarah Mitchell",    "PMO",          "Project Coordinator"),
    ("David Patel",       "Engineering",  "Lead Developer"),
    ("Emma Clarke",       "Engineering",  "Senior Developer"),
    ("Liam Foster",       "Engineering",  "Junior Developer"),
    ("Sophie Bennett",    "QA",           "QA Lead"),
    ("Oliver Hughes",     "QA",           "QA Engineer"),
    ("Charlotte Davies",  "Design",       "UX Designer"),
    ("Noah Thompson",     "Design",       "UI Designer"),
    ("Amelia Wilson",     "Business",     "Business Analyst"),
    ("George Harris",     "Business",     "Product Owner"),
    ("Isla Roberts",      "Infrastructure","DevOps Engineer"),
    ("Henry Walker",      "Security",     "Security Analyst"),
    ("Poppy White",       "Training",     "Training Manager"),
    ("Archie Martin",     "Stakeholders", "Executive Sponsor"),
]

RESOURCE_NAMES = [r[0] for r in RESOURCES]

def next_working_day(d):
    """Advance date past weekends and UK holidays."""
    d = d + timedelta(days=1)
    while d.weekday() >= 5 or d in HOLIDAY_DATES:
        d = d + timedelta(days=1)
    return d

def working_days_add(start, days):
    """Add working days to a start date (inclusive of start)."""
    d = start
    remaining = days - 1
    while remaining > 0:
        d = d + timedelta(days=1)
        if d.weekday() < 5 and d not in HOLIDAY_DATES:
            remaining -= 1
    return d

# ─── PROJECT PHASES & TASKS ──────────────────────────────────────────────────
# Each task: (type, description, duration, resource, status, pct_done, notes)
# type: "Phase", "Task", "Subtask"
# status: Not Started | In Progress | Completed | Overdue | Blocked

PHASES_DATA = [
    {
        "name": "Phase 1: Project Initiation",
        "tasks": [
            ("Task",    "Project Charter Development",        5,  "James Thornton",   "Completed",    1.0,  "Approved by Sponsor"),
            ("Subtask", "Draft project charter document",     2,  "James Thornton",   "Completed",    1.0,  ""),
            ("Subtask", "Stakeholder review & feedback",      2,  "George Harris",    "Completed",    1.0,  ""),
            ("Subtask", "Final sign-off obtained",            1,  "Archie Martin",    "Completed",    1.0,  ""),
            ("Task",    "Stakeholder Identification",         3,  "Sarah Mitchell",   "Completed",    1.0,  ""),
            ("Subtask", "Map key stakeholders",               1,  "Amelia Wilson",    "Completed",    1.0,  ""),
            ("Subtask", "Define RACI matrix",                 2,  "James Thornton",   "Completed",    1.0,  ""),
            ("Task",    "Kick-off Meeting",                   1,  "James Thornton",   "Completed",    1.0,  "All stakeholders attended"),
            ("Task",    "Project Governance Setup",           2,  "James Thornton",   "Completed",    1.0,  ""),
        ]
    },
    {
        "name": "Phase 2: Requirements Gathering",
        "tasks": [
            ("Task",    "Business Requirements Workshop",     3,  "Amelia Wilson",    "Completed",    1.0,  ""),
            ("Subtask", "Facilitated workshop session",       1,  "Amelia Wilson",    "Completed",    1.0,  ""),
            ("Subtask", "Document functional requirements",   2,  "Amelia Wilson",    "Completed",    1.0,  ""),
            ("Task",    "Technical Requirements Review",      4,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "Architecture feasibility analysis",  2,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "Non-functional requirements doc",    2,  "Emma Clarke",      "Completed",    1.0,  ""),
            ("Task",    "Requirements Sign-off Document",     2,  "George Harris",    "Completed",    1.0,  "Approved"),
            ("Task",    "Gap Analysis Report",                3,  "Amelia Wilson",    "Completed",    1.0,  ""),
        ]
    },
    {
        "name": "Phase 3: Project Planning",
        "tasks": [
            ("Task",    "Resource Planning",                  3,  "James Thornton",   "Completed",    1.0,  ""),
            ("Subtask", "Identify resource requirements",     1,  "James Thornton",   "Completed",    1.0,  ""),
            ("Subtask", "Confirm resource availability",      2,  "Sarah Mitchell",   "Completed",    1.0,  ""),
            ("Task",    "Risk Assessment & Register",         4,  "James Thornton",   "Completed",    1.0,  ""),
            ("Subtask", "Risk identification workshops",      2,  "Amelia Wilson",    "Completed",    1.0,  ""),
            ("Subtask", "Risk scoring & mitigation plans",    2,  "James Thornton",   "Completed",    1.0,  ""),
            ("Task",    "Project Schedule Development",       5,  "James Thornton",   "Completed",    1.0,  "Baselined"),
            ("Task",    "Budget Planning & Approval",         3,  "Archie Martin",    "Completed",    1.0,  "Budget approved £250k"),
            ("Task",    "Communication Plan",                 2,  "Sarah Mitchell",   "Completed",    1.0,  ""),
            ("Task",    "Quality Management Plan",            2,  "Sophie Bennett",   "Completed",    1.0,  ""),
        ]
    },
    {
        "name": "Phase 4: Design",
        "tasks": [
            ("Task",    "System Architecture Design",         8,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "High-level architecture diagram",    3,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "Component design specification",     3,  "Emma Clarke",      "Completed",    1.0,  ""),
            ("Subtask", "API design documentation",           2,  "David Patel",      "Completed",    1.0,  ""),
            ("Task",    "UX / UI Design",                     10, "Charlotte Davies", "Completed",    1.0,  ""),
            ("Subtask", "User journey mapping",               3,  "Charlotte Davies", "Completed",    1.0,  ""),
            ("Subtask", "Wireframe development",              4,  "Charlotte Davies", "Completed",    1.0,  ""),
            ("Subtask", "High-fidelity mockups",              3,  "Noah Thompson",    "Completed",    1.0,  ""),
            ("Task",    "Database Schema Design",             5,  "Emma Clarke",      "Completed",    1.0,  ""),
            ("Task",    "Design Review & Approval",           2,  "George Harris",    "Completed",    1.0,  ""),
        ]
    },
    {
        "name": "Phase 5: Development – Core",
        "tasks": [
            ("Task",    "Development Environment Setup",      3,  "Isla Roberts",     "Completed",    1.0,  "CI/CD pipelines configured"),
            ("Subtask", "Repository & branching strategy",    1,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "Dev/staging/prod environments",      2,  "Isla Roberts",     "Completed",    1.0,  ""),
            ("Task",    "Core Authentication Module",         8,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "User login & registration",          4,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "Role-based access control",          4,  "Emma Clarke",      "Completed",    1.0,  ""),
            ("Task",    "Database Implementation",            6,  "Emma Clarke",      "Completed",    1.0,  ""),
            ("Task",    "Core API Layer Development",         10, "David Patel",      "In Progress",  0.7,  "On track"),
            ("Subtask", "REST endpoints – CRUD operations",   5,  "David Patel",      "Completed",    1.0,  ""),
            ("Subtask", "Error handling & logging",           3,  "Emma Clarke",      "In Progress",  0.5,  ""),
            ("Subtask", "Unit test coverage",                 2,  "Liam Foster",      "Not Started",  0.0,  "Awaiting API completion"),
            ("Task",    "Code Review – Core Modules",         3,  "David Patel",      "Not Started",  0.0,  ""),
        ]
    },
    {
        "name": "Phase 6: Development – Features",
        "tasks": [
            ("Task",    "Feature A: Dashboard Module",        10, "Emma Clarke",      "In Progress",  0.6,  ""),
            ("Subtask", "Dashboard data aggregation",         4,  "Emma Clarke",      "In Progress",  0.8,  ""),
            ("Subtask", "Dashboard UI components",            4,  "Noah Thompson",    "In Progress",  0.5,  ""),
            ("Subtask", "Dashboard real-time updates",        2,  "David Patel",      "Not Started",  0.0,  ""),
            ("Task",    "Feature B: Reporting Engine",        8,  "Liam Foster",      "In Progress",  0.3,  "Slight delay – resource conflict"),
            ("Subtask", "Report template framework",          4,  "Liam Foster",      "In Progress",  0.5,  ""),
            ("Subtask", "PDF/Excel export functionality",     4,  "Liam Foster",      "Not Started",  0.0,  ""),
            ("Task",    "Feature C: Notifications System",    6,  "Emma Clarke",      "Not Started",  0.0,  ""),
            ("Subtask", "Email notification templates",       3,  "Emma Clarke",      "Not Started",  0.0,  ""),
            ("Subtask", "In-app notification centre",         3,  "Liam Foster",      "Not Started",  0.0,  ""),
            ("Task",    "Feature D: User Management",         5,  "David Patel",      "Not Started",  0.0,  ""),
            ("Task",    "Integration Testing – Features",     5,  "Sophie Bennett",   "Not Started",  0.0,  ""),
        ]
    },
    {
        "name": "Phase 7: Quality Assurance",
        "tasks": [
            ("Task",    "Test Plan Development",              4,  "Sophie Bennett",   "In Progress",  0.5,  ""),
            ("Subtask", "Define test scenarios & cases",      2,  "Sophie Bennett",   "Completed",    1.0,  ""),
            ("Subtask", "Test environment setup",             2,  "Oliver Hughes",    "In Progress",  0.5,  ""),
            ("Task",    "System Testing",                     10, "Oliver Hughes",    "Not Started",  0.0,  "Dependent on Dev completion"),
            ("Subtask", "Functional system testing",          5,  "Oliver Hughes",    "Not Started",  0.0,  ""),
            ("Subtask", "Regression test suite",              3,  "Sophie Bennett",   "Not Started",  0.0,  ""),
            ("Subtask", "Accessibility testing",              2,  "Oliver Hughes",    "Not Started",  0.0,  ""),
            ("Task",    "User Acceptance Testing",            10, "George Harris",    "Not Started",  0.0,  "Stakeholders to participate"),
            ("Subtask", "UAT test scripts preparation",       3,  "Amelia Wilson",    "Not Started",  0.0,  ""),
            ("Subtask", "UAT execution",                      5,  "George Harris",    "Not Started",  0.0,  ""),
            ("Subtask", "UAT sign-off",                       2,  "George Harris",    "Not Started",  0.0,  ""),
            ("Task",    "Bug Triage & Resolution",            8,  "David Patel",      "Not Started",  0.0,  ""),
            ("Task",    "Performance Testing",                5,  "Oliver Hughes",    "Not Started",  0.0,  "Load & stress testing"),
        ]
    },
    {
        "name": "Phase 8: Security & Integration",
        "tasks": [
            ("Task",    "Security Audit",                     5,  "Henry Walker",     "Not Started",  0.0,  "External pen test booked"),
            ("Subtask", "Vulnerability scanning",             2,  "Henry Walker",     "Not Started",  0.0,  ""),
            ("Subtask", "Penetration testing",                2,  "Henry Walker",     "Not Started",  0.0,  ""),
            ("Subtask", "Security remediation",               1,  "David Patel",      "Not Started",  0.0,  ""),
            ("Task",    "Third-party Integration",            6,  "David Patel",      "Not Started",  0.0,  "CRM & ERP connectors"),
            ("Subtask", "CRM API integration",                3,  "David Patel",      "Not Started",  0.0,  ""),
            ("Subtask", "ERP data sync",                      3,  "Emma Clarke",      "Not Started",  0.0,  ""),
            ("Task",    "End-to-End Integration Testing",     5,  "Sophie Bennett",   "Not Started",  0.0,  ""),
            ("Task",    "Data Migration & Validation",        5,  "Amelia Wilson",    "Not Started",  0.0,  "Legacy data migration plan"),
            ("Subtask", "Data mapping specification",         2,  "Amelia Wilson",    "Not Started",  0.0,  ""),
            ("Subtask", "Migration dry-run & validation",     3,  "Isla Roberts",     "Not Started",  0.0,  ""),
        ]
    },
    {
        "name": "Phase 9: Infrastructure & Deployment Prep",
        "tasks": [
            ("Task",    "Production Infrastructure Setup",   8,  "Isla Roberts",     "Not Started",  0.0,  "Cloud provisioning required"),
            ("Subtask", "Cloud environment provisioning",     3,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "Network & firewall configuration",   2,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "Monitoring & alerting setup",        3,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Task",    "Deployment Pipeline Automation",     5,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Task",    "Disaster Recovery & Backup Plan",    4,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "Backup procedures documentation",    2,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "DR runbook creation",                2,  "David Patel",      "Not Started",  0.0,  ""),
            ("Task",    "Rollback Plan Development",          2,  "James Thornton",   "Not Started",  0.0,  ""),
            ("Task",    "Go-Live Deployment Checklist",       3,  "James Thornton",   "Not Started",  0.0,  ""),
        ]
    },
    {
        "name": "Phase 10: Training",
        "tasks": [
            ("Task",    "Training Materials Development",     8,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Subtask", "End-user training manuals",          3,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Subtask", "System admin guide",                 2,  "David Patel",      "Not Started",  0.0,  ""),
            ("Subtask", "Video tutorial creation",            3,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Task",    "IT/Admin Team Training",             3,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Subtask", "System administration training",     2,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "Helpdesk procedure training",        1,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Task",    "End-User Training Rollout",          5,  "Poppy White",      "Not Started",  0.0,  "3 sessions planned"),
            ("Subtask", "Training session 1 – Finance",       1,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Subtask", "Training session 2 – Operations",    1,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Subtask", "Training session 3 – Management",    1,  "Poppy White",      "Not Started",  0.0,  ""),
            ("Task",    "Training Completion & Sign-off",     2,  "George Harris",    "Not Started",  0.0,  ""),
        ]
    },
    {
        "name": "Phase 11: Go-Live",
        "tasks": [
            ("Task",    "Go / No-Go Decision Meeting",        1,  "Archie Martin",    "Not Started",  0.0,  "All leads to attend"),
            ("Subtask", "Readiness assessment review",        0.5,"James Thornton",   "Not Started",  0.0,  ""),
            ("Subtask", "Formal go/no-go sign-off",           0.5,"Archie Martin",    "Not Started",  0.0,  ""),
            ("Task",    "Production Deployment",              2,  "Isla Roberts",     "Not Started",  0.0,  "Scheduled weekend deployment"),
            ("Subtask", "Pre-deployment checks",              0.5,"Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "Deployment execution",               1,  "Isla Roberts",     "Not Started",  0.0,  ""),
            ("Subtask", "Smoke testing post-deployment",      0.5,"Sophie Bennett",   "Not Started",  0.0,  ""),
            ("Task",    "Hypercare Support Period",           10, "David Patel",      "Not Started",  0.0,  "2-week hypercare window"),
            ("Subtask", "24/7 on-call support rota",          5,  "David Patel",      "Not Started",  0.0,  ""),
            ("Subtask", "Issue triage & resolution",          5,  "Emma Clarke",      "Not Started",  0.0,  ""),
            ("Task",    "Post Go-Live Review",                2,  "James Thornton",   "Not Started",  0.0,  ""),
        ]
    },
    {
        "name": "Phase 12: Project Closure",
        "tasks": [
            ("Task",    "Lessons Learned Workshop",           2,  "James Thornton",   "Not Started",  0.0,  "All team members to contribute"),
            ("Subtask", "Facilitate lessons learned session", 1,  "James Thornton",   "Not Started",  0.0,  ""),
            ("Subtask", "Document and distribute outcomes",   1,  "Sarah Mitchell",   "Not Started",  0.0,  ""),
            ("Task",    "Final Project Documentation",        5,  "Sarah Mitchell",   "Not Started",  0.0,  ""),
            ("Subtask", "As-built documentation",             2,  "David Patel",      "Not Started",  0.0,  ""),
            ("Subtask", "Project archive preparation",        2,  "Sarah Mitchell",   "Not Started",  0.0,  ""),
            ("Subtask", "Handover pack assembly",             1,  "James Thornton",   "Not Started",  0.0,  ""),
            ("Task",    "Project Formal Handover",            1,  "Archie Martin",    "Not Started",  0.0,  ""),
            ("Task",    "Budget Reconciliation Report",       3,  "James Thornton",   "Not Started",  0.0,  ""),
            ("Task",    "Project Closure Report Sign-off",    2,  "Archie Martin",    "Not Started",  0.0,  ""),
            ("Task",    "Benefits Realisation Plan Setup",    3,  "George Harris",    "Not Started",  0.0,  "Handed to BAU team"),
        ]
    },
]

# ─── BUILD FLAT TASK LIST WITH IDs, DATES ────────────────────────────────────
def build_task_list():
    """Compute task IDs and dates sequentially."""
    tasks = []
    task_id = 0
    project_start = date(2025, 3, 3)  # First working day in March 2025
    current_date = project_start

    # Ensure project_start is a working day
    while current_date.weekday() >= 5 or current_date in HOLIDAY_DATES:
        current_date += timedelta(days=1)

    phase_start_date = current_date

    for phase in PHASES_DATA:
        phase_start = phase_start_date
        phase_task_count = 0

        for row in phase["tasks"]:
            task_id += 1
            ttype, desc, dur, resource, status, pct, notes = row
            dur = max(1, int(dur))

            # Start = current working date
            start = current_date
            # End = start + duration in working days
            end = working_days_add(start, dur)

            tasks.append({
                "id": task_id,
                "type": ttype,
                "phase": phase["name"],
                "desc": desc,
                "start": start,
                "end": end,
                "duration": dur,
                "resource": resource,
                "pct": pct,
                "status": status,
                "notes": notes,
            })

            # Subtasks share parent date; Tasks advance date
            if ttype == "Task":
                current_date = next_working_day(end)
            elif ttype == "Subtask":
                pass  # Subtasks overlap within task

            phase_task_count += 1

        # Move to next working day after phase ends
        if tasks:
            last_end = max(t["end"] for t in tasks if t["phase"] == phase["name"])
            phase_start_date = next_working_day(last_end)
            current_date = phase_start_date

    return tasks, project_start

# ─── PROJECT PLAN SHEET ──────────────────────────────────────────────────────
def setup_project_plan(wb, tasks, project_start):
    ws = wb.active
    ws.title = "Project Plan"
    ws.sheet_view.showGridLines = True
    ws.sheet_properties.outlinePr.summaryBelow = False  # Expand buttons at top
    ws.freeze_panes = "B12"

    # Column widths
    col_widths = {
        "A": 8,   # Task ID
        "B": 12,  # Task Type
        "C": 30,  # Phase
        "D": 45,  # Description
        "E": 14,  # Planned Start
        "F": 14,  # Planned End
        "G": 22,  # Resource
        "H": 10,  # % Done
        "I": 14,  # Status
        "J": 40,  # Notes / Blockers
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── PROJECT SUMMARY SECTION (rows 1-9) ────────────────────────────────
    # Row 1: Big title
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "PROJECT MANAGEMENT PLAN"
    c.font = Font(name="Calibri", bold=True, size=20, color=C_WHITE)
    c.fill = fill(C_SUMM_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Project name field
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "Project Name:"
    c.font = Font(name="Calibri", bold=True, size=12, color=C_WHITE)
    c.fill = fill(C_SUMM_ACCNT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("E2:J2")
    c = ws["E2"]
    c.value = "Digital Transformation Programme"
    c.font = Font(name="Calibri", bold=False, size=12, color="000000")
    c.fill = fill(C_YELLOW_HDR)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Row 3: Date summary — fixed non-overlapping ranges
    ws.row_dimensions[3].height = 24
    date_sections = [
        ("A3", "B3", "Project Start:", "C3", "D3", project_start.strftime("%d/%m/%Y")),
        ("E3", "F3", "Project End:",   "G3", "H3", '=TEXT(MAX(F12:F500),"DD/MM/YYYY")'),
        ("I3", "I3", "Last Updated:",  "J3", "J3",  date.today().strftime("%d/%m/%Y")),
    ]
    for lbl_start, lbl_end, label, val_start, val_end, val in date_sections:
        ws.merge_cells(f"{lbl_start}:{lbl_end}")
        c = ws[lbl_start]
        c.value = label
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(C_SUMM_ACCNT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"{val_start}:{val_end}")
        c2 = ws[val_start]
        c2.value = val
        c2.font = Font(name="Calibri", size=10)
        c2.fill = fill(C_YELLOW_HDR)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Row 4: KPI labels
    ws.row_dimensions[4].height = 22
    kpi_labels = ["Total Tasks", "Completed", "In Progress", "Overdue", "Blocked", "% Complete"]
    kpi_cols = ["A", "B", "C", "D", "E", "F"]
    for col, label in zip(kpi_cols, kpi_labels):
        c = ws[f"{col}4"]
        c.value = label
        c.font = Font(name="Calibri", bold=True, size=9, color=C_WHITE)
        c.fill = fill(C_SUMM_ACCNT)
        c.alignment = center()

    # Row 5: KPI values using COUNTIF formulas
    ws.row_dimensions[5].height = 28
    kpi_formulas = [
        ('=COUNTIF(B12:B500,"Task")+COUNTIF(B12:B500,"Subtask")', "000000", C_LIGHT_BLUE),
        ('=COUNTIF(I12:I500,"Completed")', C_WHITE, C_GREEN),
        ('=COUNTIF(I12:I500,"In Progress")', "000000", C_AMBER),
        ('=COUNTIF(I12:I500,"Overdue")', C_WHITE, C_RED),
        ('=COUNTIF(I12:I500,"Blocked")', C_WHITE, C_PURPLE),
        ('=IFERROR(COUNTIF(I12:I500,"Completed")/(COUNTIF(B12:B500,"Task")+COUNTIF(B12:B500,"Subtask")),0)', C_WHITE, C_NAVY),
    ]
    for col, (formula, fcolor, bg) in zip(kpi_cols, kpi_formulas):
        c = ws[f"{col}5"]
        c.value = formula
        c.font = Font(name="Calibri", bold=True, size=14, color=fcolor)
        c.fill = fill(bg)
        c.alignment = center()
        if col == "F":
            c.number_format = "0.0%"

    # Row 6: Merge unused KPI area / status breakdown label
    ws.row_dimensions[6].height = 18
    ws.merge_cells("G4:J5")
    c = ws["G4"]
    c.value = "STATUS SUMMARY"
    c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = center()

    # Row 7: Status mini-legend
    ws.row_dimensions[7].height = 20
    status_legend = [
        ("A", "● Not Started",  C_GREY_NS,  "FFFFFF"),
        ("C", "● In Progress",  C_AMBER,    "FFFFFF"),
        ("E", "● Completed",    C_GREEN,    "FFFFFF"),
        ("G", "● Overdue",      C_RED,      "FFFFFF"),
        ("I", "● Blocked",      C_PURPLE,   "FFFFFF"),
    ]
    for col, label, bg, fg in status_legend:
        ws.merge_cells(f"{col}7:{chr(ord(col)+1)}7")
        c = ws[f"{col}7"]
        c.value = label
        c.font = Font(name="Calibri", bold=True, size=9, color=fg)
        c.fill = fill(bg)
        c.alignment = center()

    # Rows 8-9: Button placeholders (VBA will assign macros)
    ws.row_dimensions[8].height = 10
    ws.row_dimensions[9].height = 32
    ws.row_dimensions[10].height = 10

    # Draw button labels as cells (user will add shapes over these)
    # Each tuple: (start_col, end_col, label, bg_colour, fg_colour)
    btn_configs = [
        ("B", "D", "EXPORT GANTT CHART",  C_NAVY,      C_WHITE),
        ("E", "G", "EXPORT DASHBOARD",     C_SUMM_ACCNT, C_WHITE),
        ("H", "J", "EXPORT RAG REPORT",    "375623",    C_WHITE),
    ]
    for start_col, end_col, label, bg, fg in btn_configs:
        ws.merge_cells(f"{start_col}9:{end_col}9")
        c = ws[f"{start_col}9"]
        c.value = label
        c.font = Font(name="Calibri", bold=True, size=11, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        btn_border = Side(style="medium", color="FFFFFF")
        c.border = Border(left=btn_border, right=btn_border, top=btn_border, bottom=btn_border)

    # ── COLUMN HEADERS (row 11) ────────────────────────────────────────────
    ws.row_dimensions[11].height = 28
    headers = [
        ("A11", "ID"),
        ("B11", "Task Type"),
        ("C11", "Phase"),
        ("D11", "Task / Subtask Description"),
        ("E11", "Planned Start"),
        ("F11", "Planned End"),
        ("G11", "Resource"),
        ("H11", "% Done"),
        ("I11", "Status"),
        ("J11", "Notes / Blockers"),
    ]
    for cell_ref, header in headers:
        c = ws[cell_ref]
        c.value = header
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(C_NAVY)
        c.alignment = center()
        c.border = border_all()

    # ── TASK DATA (row 12+) ────────────────────────────────────────────────
    HEADER_ROW = 11
    DATA_START = 12

    row_num = DATA_START
    phase_row_map = {}  # phase_name -> list of data rows

    current_phase = None
    phase_start_row = DATA_START

    for i, task in enumerate(tasks):
        # Insert phase header row when phase changes
        if task["phase"] != current_phase:
            if current_phase is not None:
                # Group the rows of previous phase
                phase_end_row = row_num - 1
                if phase_end_row >= phase_start_row:
                    ws.row_dimensions.group(phase_start_row, phase_end_row, outline_level=1, hidden=False)

            current_phase = task["phase"]
            phase_start_row = row_num + 1  # Rows after header

            # Phase header row
            ws.row_dimensions[row_num].height = 24
            ws[f"A{row_num}"].value = ""
            ws[f"B{row_num}"].value = "Phase"
            ws[f"C{row_num}"].value = current_phase
            ws[f"D{row_num}"].value = current_phase

            # Phase start and end (formula-driven)
            phase_idx = [p["name"] for p in PHASES_DATA].index(current_phase)
            # We'll fill with actual values
            phase_tasks = [t for t in tasks if t["phase"] == current_phase and t["type"] == "Task"]
            if phase_tasks:
                ws[f"E{row_num}"].value = phase_tasks[0]["start"]
                ws[f"F{row_num}"].value = max(t["end"] for t in phase_tasks)
            ws[f"G{row_num}"].value = ""
            ws[f"H{row_num}"].value = ""
            ws[f"I{row_num}"].value = ""
            ws[f"J{row_num}"].value = ""

            # Style phase header
            for col in "ABCDEFGHIJ":
                c = ws[f"{col}{row_num}"]
                c.fill = fill(C_NAVY)
                c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                thin = Side(style="medium", color="FFFFFF")
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws[f"C{row_num}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

            # Date formatting for phase header
            ws[f"E{row_num}"].number_format = "DD/MM/YYYY"
            ws[f"F{row_num}"].number_format = "DD/MM/YYYY"

            row_num += 1

        # Task / Subtask row
        ws.row_dimensions[row_num].height = 20

        is_subtask = task["type"] == "Subtask"
        row_bg = C_SUBTASK_BG if is_subtask else C_WHITE
        indent = 2 if is_subtask else 1

        ws[f"A{row_num}"].value = task["id"]
        ws[f"B{row_num}"].value = task["type"]
        ws[f"C{row_num}"].value = task["phase"]
        ws[f"D{row_num}"].value = ("    ↳  " if is_subtask else "") + task["desc"]
        ws[f"E{row_num}"].value = task["start"]
        ws[f"F{row_num}"].value = task["end"]
        ws[f"G{row_num}"].value = task["resource"]
        ws[f"H{row_num}"].value = task["pct"]
        ws[f"I{row_num}"].value = task["status"]
        ws[f"J{row_num}"].value = task["notes"]

        # Number formats
        ws[f"E{row_num}"].number_format = "DD/MM/YYYY"
        ws[f"F{row_num}"].number_format = "DD/MM/YYYY"
        ws[f"H{row_num}"].number_format = "0%"

        # Style
        status_bg = {
            "Completed":   (C_LIGHT_GRN, "000000"),
            "In Progress": (C_LIGHT_AMBE, "000000"),
            "Not Started": (C_WHITE,      "000000"),
            "Overdue":     (C_LIGHT_RED,  "000000"),
            "Blocked":     ("EDD9F5",     "000000"),
        }.get(task["status"], (row_bg, "000000"))

        for col in "ABCDEFGHIJ":
            c = ws[f"{col}{row_num}"]
            if is_subtask:
                c.fill = fill(C_SUBTASK_BG)
            else:
                c.fill = fill(C_WHITE)
            c.font = Font(name="Calibri", size=9 if is_subtask else 10,
                          italic=is_subtask, color="000000")
            c.border = border_all()

        ws[f"D{row_num}"].alignment = Alignment(horizontal="left", vertical="center",
                                                  indent=indent, wrap_text=True)
        ws[f"J{row_num}"].alignment = Alignment(horizontal="left", vertical="center",
                                                  wrap_text=True)

        # Status cell coloring
        status_color_map = {
            "Not Started": (C_GREY_NS, C_WHITE),
            "In Progress": (C_AMBER,   "000000"),
            "Completed":   (C_GREEN,   C_WHITE),
            "Overdue":     (C_RED,     C_WHITE),
            "Blocked":     (C_PURPLE,  C_WHITE),
        }
        sc = ws[f"I{row_num}"]
        sc_bg, sc_fg = status_color_map.get(task["status"], ("FFFFFF", "000000"))
        sc.fill = fill(sc_bg)
        sc.font = Font(name="Calibri", size=9, bold=True, color=sc_fg)
        sc.alignment = center()

        # Task type cell
        tc = ws[f"B{row_num}"]
        type_colors = {"Task": (C_NAVY_MID, C_WHITE), "Subtask": ("7F7F7F", C_WHITE)}
        tc_bg, tc_fg = type_colors.get(task["type"], ("FFFFFF", "000000"))
        tc.fill = fill(tc_bg)
        tc.font = Font(name="Calibri", size=9, bold=True, color=tc_fg)
        tc.alignment = center()

        # ID cell style
        ws[f"A{row_num}"].alignment = center()
        ws[f"A{row_num}"].font = Font(name="Calibri", size=9, bold=True, color=C_NAVY)

        # % Done cell
        ws[f"H{row_num}"].alignment = center()
        ws[f"H{row_num}"].font = Font(name="Calibri", size=9, bold=True)

        row_num += 1

    # Group the last phase
    if phase_start_row < row_num:
        ws.row_dimensions.group(phase_start_row, row_num - 1, outline_level=1, hidden=False)

    # ── DATA VALIDATION ───────────────────────────────────────────────────
    last_data_row = row_num - 1

    # Task Type dropdown
    dv_type = DataValidation(
        type="list",
        formula1='"Task,Subtask,Phase"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Type",
        error="Please select: Task, Subtask, or Phase"
    )
    ws.add_data_validation(dv_type)
    dv_type.add(f"B{DATA_START}:B{last_data_row + 100}")

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Overdue,Blocked"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Choose: Not Started, In Progress, Completed, Overdue, or Blocked"
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"I{DATA_START}:I{last_data_row + 100}")

    # Resource dropdown (from Resources sheet)
    dv_resource = DataValidation(
        type="list",
        formula1="=Resources!$A$2:$A$50",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Resource",
        error="Resource must exist in the Resources sheet"
    )
    ws.add_data_validation(dv_resource)
    dv_resource.add(f"G{DATA_START}:G{last_data_row + 100}")

    # % Done validation (0 to 1)
    dv_pct = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid %",
        error="Enter a value between 0 (0%) and 1 (100%), e.g. 0.5 = 50%"
    )
    ws.add_data_validation(dv_pct)
    dv_pct.add(f"H{DATA_START}:H{last_data_row + 100}")

    return row_num - 1  # last data row


# ─── RESOURCES SHEET ─────────────────────────────────────────────────────────
def setup_resources(wb):
    ws = wb["Resources"]
    ws.sheet_view.showGridLines = True

    col_widths = {"A": 25, "B": 20, "C": 22, "D": 30}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "RESOURCE REGISTER"
    c.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = center()

    headers = ["Name", "Department", "Role", "Email / Contact"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(C_SUMM_ACCNT)
        c.alignment = center()
        c.border = border_all()

    sample_emails = [
        "j.thornton@company.com", "s.mitchell@company.com", "d.patel@company.com",
        "e.clarke@company.com", "l.foster@company.com", "s.bennett@company.com",
        "o.hughes@company.com", "c.davies@company.com", "n.thompson@company.com",
        "a.wilson@company.com", "g.harris@company.com", "i.roberts@company.com",
        "h.walker@company.com", "p.white@company.com", "a.martin@company.com",
    ]

    for r_idx, (resource, email) in enumerate(zip(RESOURCES, sample_emails), 3):
        name, dept, role = resource
        row_bg = C_LIGHT_BLUE if r_idx % 2 == 0 else C_WHITE
        for col_idx, val in enumerate([name, dept, role, email], 1):
            c = ws.cell(row=r_idx, column=col_idx)
            c.value = val
            c.font = Font(name="Calibri", size=10)
            c.fill = fill(row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = border_all()

    ws.freeze_panes = "A3"


# ─── HOLIDAYS SHEET ──────────────────────────────────────────────────────────
def setup_holidays(wb):
    ws = wb["Holidays"]
    ws.sheet_view.showGridLines = True

    col_widths = {"A": 35, "B": 16, "C": 16, "D": 20, "E": 30}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "UK PUBLIC HOLIDAYS CALENDAR"
    c.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = center()

    # Info note
    ws.merge_cells("A2:E2")
    note = ws["A2"]
    note.value = ("⚠  Holiday dates in column B are used by VBA macros to auto-adjust task dates. "
                  "Leave the Resource column blank for public holidays (applies to everyone). "
                  "Add resource name for individual leave.")
    note.font = Font(name="Calibri", size=9, italic=True, color="7F4A00")
    note.fill = fill("FFF2CC")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 30

    headers = ["Holiday / Leave Name", "Start Date", "End Date", "Resource (blank = all)", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(C_SUMM_ACCNT)
        c.alignment = center()
        c.border = border_all()

    # Group holidays by year
    sorted_holidays = sorted(UK_HOLIDAYS, key=lambda x: x[1])
    current_year = None

    data_row = 4
    for name, start, end, resource in sorted_holidays:
        yr = start.year
        if yr != current_year:
            current_year = yr
            ws.row_dimensions[data_row].height = 20
            ws.merge_cells(f"A{data_row}:E{data_row}")
            yc = ws[f"A{data_row}"]
            yc.value = f"── {yr} Public Holidays ──"
            yc.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
            yc.fill = fill(C_NAVY_MID)
            yc.alignment = center()
            data_row += 1

        row_bg = "EBF3FB" if data_row % 2 == 0 else C_WHITE
        ws.cell(row=data_row, column=1).value = name
        ws.cell(row=data_row, column=2).value = start
        ws.cell(row=data_row, column=3).value = end
        ws.cell(row=data_row, column=4).value = resource or ""
        ws.cell(row=data_row, column=5).value = "England & Wales"

        for col_idx in range(1, 6):
            c = ws.cell(row=data_row, column=col_idx)
            c.font = Font(name="Calibri", size=10)
            c.fill = fill(row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = border_all()

        ws.cell(row=data_row, column=2).number_format = "DD/MM/YYYY"
        ws.cell(row=data_row, column=3).number_format = "DD/MM/YYYY"
        data_row += 1

    ws.freeze_panes = "A4"


# ─── STAKEHOLDERS SHEET ──────────────────────────────────────────────────────
def setup_stakeholders(wb):
    ws = wb["Stakeholders"]
    ws.sheet_view.showGridLines = True

    col_widths = {"A": 25, "B": 35, "C": 10, "D": 20}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "STAKEHOLDER EMAIL DISTRIBUTION LIST"
    c.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = center()

    headers = ["Name", "Email", "Type (To/CC/BCC)", "Organisation"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(C_SUMM_ACCNT)
        c.alignment = center()
        c.border = border_all()

    stakeholders_data = [
        ("Archie Martin",     "a.martin@company.com",        "To",  "Executive Team"),
        ("George Harris",     "g.harris@company.com",        "To",  "Product"),
        ("James Thornton",    "j.thornton@company.com",      "To",  "PMO"),
        ("David Patel",       "d.patel@company.com",         "CC",  "Engineering"),
        ("Sophie Bennett",    "s.bennett@company.com",       "CC",  "QA"),
        ("Sarah Mitchell",    "s.mitchell@company.com",      "CC",  "PMO"),
        ("Amelia Wilson",     "a.wilson@company.com",        "CC",  "Business"),
        ("Board Secretary",   "board@company.com",           "BCC", "Board"),
    ]

    dv_type = DataValidation(type="list", formula1='"To,CC,BCC"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add("C3:C100")

    for r_idx, (name, email, stype, org) in enumerate(stakeholders_data, 3):
        row_bg = C_LIGHT_BLUE if r_idx % 2 == 0 else C_WHITE
        for col_idx, val in enumerate([name, email, stype, org], 1):
            c = ws.cell(row=r_idx, column=col_idx)
            c.value = val
            c.font = Font(name="Calibri", size=10)
            c.fill = fill(row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = border_all()

    ws.freeze_panes = "A3"


# ─── VBA CODE ─────────────────────────────────────────────────────────────────
VBA_CODE = r'''
'=============================================================================
' PROJECT MANAGEMENT TOOL - VBA MODULE
' Version 1.0 | Import via: VBA Editor (Alt+F11) > File > Import File
'=============================================================================
Option Explicit

'─────────────────────────────────────────────────────────────────────────────
' CONSTANTS & HELPERS
'─────────────────────────────────────────────────────────────────────────────
Private Const PLAN_SHEET As String = "Project Plan"
Private Const DATA_START_ROW As Long = 12
Private Const COL_ID As Integer = 1
Private Const COL_TYPE As Integer = 2
Private Const COL_PHASE As Integer = 3
Private Const COL_DESC As Integer = 4
Private Const COL_START As Integer = 5
Private Const COL_END As Integer = 6
Private Const COL_RESOURCE As Integer = 7
Private Const COL_PCT As Integer = 8
Private Const COL_STATUS As Integer = 9
Private Const COL_NOTES As Integer = 10

'─────────────────────────────────────────────────────────────────────────────
' SETUP PROJECT UI
' Run this once after opening the workbook (or after adding new rows)
'─────────────────────────────────────────────────────────────────────────────
Sub SetupProjectUI()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)
    Application.ScreenUpdating = False

    ' Apply conditional formatting for Status column (I)
    Dim lastRow As Long
    lastRow = FindLastDataRow(ws)

    With ws.Range("I" & DATA_START_ROW & ":I" & lastRow + 500)
        .FormatConditions.Delete

        ' Completed - Green
        Dim cfComplete As FormatCondition
        Set cfComplete = .FormatConditions.Add(Type:=xlCellValue, Operator:=xlEqual, Formula1:="=""Completed""")
        cfComplete.Interior.Color = RGB(0, 176, 80)
        cfComplete.Font.Color = RGB(255, 255, 255)
        cfComplete.Font.Bold = True

        ' In Progress - Amber
        Dim cfInProg As FormatCondition
        Set cfInProg = .FormatConditions.Add(Type:=xlCellValue, Operator:=xlEqual, Formula1:="=""In Progress""")
        cfInProg.Interior.Color = RGB(255, 185, 0)
        cfInProg.Font.Color = RGB(0, 0, 0)
        cfInProg.Font.Bold = True

        ' Overdue - Red
        Dim cfOverdue As FormatCondition
        Set cfOverdue = .FormatConditions.Add(Type:=xlCellValue, Operator:=xlEqual, Formula1:="=""Overdue""")
        cfOverdue.Interior.Color = RGB(255, 0, 0)
        cfOverdue.Font.Color = RGB(255, 255, 255)
        cfOverdue.Font.Bold = True

        ' Blocked - Purple
        Dim cfBlocked As FormatCondition
        Set cfBlocked = .FormatConditions.Add(Type:=xlCellValue, Operator:=xlEqual, Formula1:="=""Blocked""")
        cfBlocked.Interior.Color = RGB(112, 48, 160)
        cfBlocked.Font.Color = RGB(255, 255, 255)
        cfBlocked.Font.Bold = True

        ' Not Started - Grey
        Dim cfNotStart As FormatCondition
        Set cfNotStart = .FormatConditions.Add(Type:=xlCellValue, Operator:=xlEqual, Formula1:="=""Not Started""")
        cfNotStart.Interior.Color = RGB(165, 165, 165)
        cfNotStart.Font.Color = RGB(255, 255, 255)
        cfNotStart.Font.Bold = True
    End With

    MsgBox "Project UI setup complete. Conditional formatting applied.", vbInformation, "Setup Complete"
    Application.ScreenUpdating = True
End Sub

'─────────────────────────────────────────────────────────────────────────────
' AUTO-RENUMBER IDs
'─────────────────────────────────────────────────────────────────────────────
Sub RenumberIDs()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)
    Application.ScreenUpdating = False
    Application.EnableEvents = False

    Dim i As Long
    Dim taskID As Long
    taskID = 0

    Dim lastRow As Long
    lastRow = FindLastDataRow(ws)

    For i = DATA_START_ROW To lastRow
        Dim cellType As String
        cellType = ws.Cells(i, COL_TYPE).Value
        If cellType = "Task" Or cellType = "Subtask" Then
            taskID = taskID + 1
            ws.Cells(i, COL_ID).Value = taskID
        ElseIf cellType = "Phase" Then
            ws.Cells(i, COL_ID).Value = ""
        End If
    Next i

    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub

'─────────────────────────────────────────────────────────────────────────────
' IS WORKING DAY (skips weekends and UK holidays)
'─────────────────────────────────────────────────────────────────────────────
Function IsWorkingDay(dt As Date) As Boolean
    If Weekday(dt, vbMonday) >= 6 Then
        IsWorkingDay = False
        Exit Function
    End If

    Dim wsH As Worksheet
    On Error Resume Next
    Set wsH = ThisWorkbook.Sheets("Holidays")
    On Error GoTo 0

    If wsH Is Nothing Then
        IsWorkingDay = True
        Exit Function
    End If

    Dim lastHRow As Long
    lastHRow = wsH.Cells(wsH.Rows.Count, 2).End(xlUp).Row

    Dim r As Long
    For r = 4 To lastHRow
        Dim hStart As Date
        Dim hEnd As Date
        If wsH.Cells(r, 2).Value <> "" And IsDate(wsH.Cells(r, 2).Value) Then
            hStart = CDate(wsH.Cells(r, 2).Value)
            If wsH.Cells(r, 3).Value <> "" Then
                hEnd = CDate(wsH.Cells(r, 3).Value)
            Else
                hEnd = hStart
            End If
            If dt >= hStart And dt <= hEnd Then
                IsWorkingDay = False
                Exit Function
            End If
        End If
    Next r

    IsWorkingDay = True
End Function

Function NextWorkingDay(dt As Date) As Date
    Dim d As Date
    d = dt + 1
    Do While Not IsWorkingDay(d)
        d = d + 1
    Loop
    NextWorkingDay = d
End Function

Function AddWorkingDays(startDate As Date, days As Long) As Date
    Dim d As Date
    d = startDate
    Dim remaining As Long
    remaining = days - 1
    Do While remaining > 0
        d = d + 1
        If IsWorkingDay(d) Then remaining = remaining - 1
    Loop
    AddWorkingDays = d
End Function

'─────────────────────────────────────────────────────────────────────────────
' AUTO-ADJUST DATES FOR HOLIDAYS
'─────────────────────────────────────────────────────────────────────────────
Sub AdjustDatesForHolidays()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)
    Application.ScreenUpdating = False
    Application.EnableEvents = False

    Dim lastRow As Long
    lastRow = FindLastDataRow(ws)
    Dim i As Long

    For i = DATA_START_ROW To lastRow
        Dim cellType As String
        cellType = ws.Cells(i, COL_TYPE).Value
        If cellType = "Task" Or cellType = "Subtask" Then
            If IsDate(ws.Cells(i, COL_START).Value) Then
                Dim startDt As Date
                startDt = CDate(ws.Cells(i, COL_START).Value)
                ' Move start to next working day if it falls on holiday
                If Not IsWorkingDay(startDt) Then
                    Do While Not IsWorkingDay(startDt)
                        startDt = startDt + 1
                    Loop
                    ws.Cells(i, COL_START).Value = startDt
                End If
                ' Recalculate end date if duration exists (end = start + duration - 1 working days)
                ' End date cell is protected but we update it here
                If IsDate(ws.Cells(i, COL_END).Value) Then
                    Dim endDt As Date
                    endDt = CDate(ws.Cells(i, COL_END).Value)
                    If Not IsWorkingDay(endDt) Then
                        Do While Not IsWorkingDay(endDt)
                            endDt = endDt + 1
                        Loop
                        ws.Cells(i, COL_END).Value = endDt
                    End If
                End If
            End If
        End If
    Next i

    Application.EnableEvents = True
    Application.ScreenUpdating = True
    MsgBox "Dates adjusted. All task start/end dates now fall on working days.", vbInformation, "Date Adjustment Complete"
End Sub

'─────────────────────────────────────────────────────────────────────────────
' HELPER: Find last row with data
'─────────────────────────────────────────────────────────────────────────────
Function FindLastDataRow(ws As Worksheet) As Long
    FindLastDataRow = ws.Cells(ws.Rows.Count, COL_DESC).End(xlUp).Row
End Function

'─────────────────────────────────────────────────────────────────────────────
' COLLECT ALL TASKS INTO ARRAY FOR EXPORT
'─────────────────────────────────────────────────────────────────────────────
Function CollectTasks() As Variant
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)

    Dim lastRow As Long
    lastRow = FindLastDataRow(ws)

    Dim tasks() As Variant
    Dim taskCount As Long
    taskCount = 0

    Dim i As Long
    For i = DATA_START_ROW To lastRow
        If ws.Cells(i, COL_TYPE).Value <> "" Then
            taskCount = taskCount + 1
        End If
    Next i

    ReDim tasks(1 To taskCount, 1 To 10)
    Dim idx As Long
    idx = 0

    For i = DATA_START_ROW To lastRow
        If ws.Cells(i, COL_TYPE).Value <> "" Then
            idx = idx + 1
            tasks(idx, 1) = ws.Cells(i, COL_ID).Value        ' ID
            tasks(idx, 2) = ws.Cells(i, COL_TYPE).Value      ' Type
            tasks(idx, 3) = ws.Cells(i, COL_PHASE).Value     ' Phase
            tasks(idx, 4) = ws.Cells(i, COL_DESC).Value      ' Description
            If IsDate(ws.Cells(i, COL_START).Value) Then
                tasks(idx, 5) = Format(ws.Cells(i, COL_START).Value, "DD/MM/YYYY")
            Else
                tasks(idx, 5) = ""
            End If
            If IsDate(ws.Cells(i, COL_END).Value) Then
                tasks(idx, 6) = Format(ws.Cells(i, COL_END).Value, "DD/MM/YYYY")
            Else
                tasks(idx, 6) = ""
            End If
            tasks(idx, 7) = ws.Cells(i, COL_RESOURCE).Value  ' Resource
            tasks(idx, 8) = ws.Cells(i, COL_PCT).Value       ' % Done
            tasks(idx, 9) = ws.Cells(i, COL_STATUS).Value    ' Status
            tasks(idx, 10) = ws.Cells(i, COL_NOTES).Value    ' Notes
        End If
    Next i

    CollectTasks = tasks
End Function

'─────────────────────────────────────────────────────────────────────────────
' GET WORKBOOK FOLDER PATH
'─────────────────────────────────────────────────────────────────────────────
Function GetSavePath(filename As String) As String
    Dim wbPath As String
    wbPath = ThisWorkbook.Path
    If wbPath = "" Then
        wbPath = Environ("USERPROFILE") & "\Desktop"
    End If
    GetSavePath = wbPath & "\" & filename
End Function

'─────────────────────────────────────────────────────────────────────────────
' WRITE STRING TO FILE
'─────────────────────────────────────────────────────────────────────────────
Sub WriteFile(filePath As String, content As String)
    Dim fileNum As Integer
    fileNum = FreeFile()
    Open filePath For Output As #fileNum
    Print #fileNum, content
    Close #fileNum
End Sub

'─────────────────────────────────────────────────────────────────────────────
' OPEN FILE IN DEFAULT BROWSER
'─────────────────────────────────────────────────────────────────────────────
Sub OpenInBrowser(filePath As String)
    Shell "cmd /c start """" """ & filePath & """", vbHide
End Sub

'=============================================================================
' EXPORT 1: INTERACTIVE GANTT CHART
'=============================================================================
Sub ExportGanttChart()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)

    Dim projectName As String
    projectName = ws.Range("E2").Value
    If projectName = "" Then projectName = "Project Gantt Chart"

    Dim tasks As Variant
    tasks = CollectTasks()

    If Not IsArray(tasks) Then
        MsgBox "No task data found.", vbExclamation
        Exit Sub
    End If

    ' Build tasks JSON
    Dim jsonTasks As String
    jsonTasks = "["
    Dim i As Long
    For i = 1 To UBound(tasks, 1)
        If i > 1 Then jsonTasks = jsonTasks & ","
        Dim pct As Double
        If IsNumeric(tasks(i, 8)) Then pct = tasks(i, 8) Else pct = 0
        Dim startStr As String, endStr As String
        startStr = tasks(i, 5)
        endStr = tasks(i, 6)
        jsonTasks = jsonTasks & "{"
        jsonTasks = jsonTasks & """id"":" & IIf(tasks(i, 1) = "", "0", tasks(i, 1)) & ","
        jsonTasks = jsonTasks & """type"":""" & tasks(i, 2) & ""","
        jsonTasks = jsonTasks & """phase"":""" & EscapeJson(CStr(tasks(i, 3))) & ""","
        jsonTasks = jsonTasks & """desc"":""" & EscapeJson(CStr(tasks(i, 4))) & ""","
        jsonTasks = jsonTasks & """start"":""" & startStr & ""","
        jsonTasks = jsonTasks & """end"":""" & endStr & ""","
        jsonTasks = jsonTasks & """resource"":""" & EscapeJson(CStr(tasks(i, 7))) & ""","
        jsonTasks = jsonTasks & """pct"":" & pct & ","
        jsonTasks = jsonTasks & """status"":""" & EscapeJson(CStr(tasks(i, 9))) & ""","
        jsonTasks = jsonTasks & """notes"":""" & EscapeJson(CStr(tasks(i, 10))) & """"
        jsonTasks = jsonTasks & "}"
    Next i
    jsonTasks = jsonTasks & "]"

    Dim html As String
    html = BuildGanttHTML(projectName, jsonTasks)

    Dim savePath As String
    savePath = GetSavePath("ProjectGanttChart.html")
    WriteFile savePath, html
    OpenInBrowser savePath
    MsgBox "Gantt Chart exported to:" & vbCrLf & savePath, vbInformation, "Export Complete"
End Sub

Function EscapeJson(s As String) As String
    s = Replace(s, "\", "\\")
    s = Replace(s, """", "\""")
    s = Replace(s, Chr(10), "\n")
    s = Replace(s, Chr(13), "")
    EscapeJson = s
End Function

Function BuildGanttHTML(projectName As String, jsonTasks As String) As String
    Dim h As String
    h = "<!DOCTYPE html><html lang=""en""><head><meta charset=""UTF-8"">"
    h = h & "<meta name=""viewport"" content=""width=device-width,initial-scale=1"">"
    h = h & "<title>Gantt Chart - " & projectName & "</title>"
    h = h & "<style>"
    h = h & "*{box-sizing:border-box;margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif}"
    h = h & "body{background:#0d1b2a;color:#e0e8f0;min-height:100vh}"
    h = h & "header{background:linear-gradient(135deg,#1f4e79,#2e75b6);padding:20px 30px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 12px rgba(0,0,0,0.4)}"
    h = h & "header h1{font-size:1.6rem;font-weight:700;letter-spacing:.5px}"
    h = h & "header .meta{font-size:.85rem;opacity:.8}"
    h = h & ".controls{background:#152030;padding:14px 30px;display:flex;gap:12px;align-items:center;flex-wrap:wrap;border-bottom:1px solid #243a50}"
    h = h & ".controls input,.controls select{background:#1e3248;border:1px solid #2e75b6;color:#e0e8f0;padding:6px 12px;border-radius:6px;font-size:.85rem}"
    h = h & ".controls label{font-size:.85rem;color:#93b5d0}"
    h = h & ".legend{display:flex;gap:12px;margin-left:auto;align-items:center;flex-wrap:wrap}"
    h = h & ".leg{display:flex;align-items:center;gap:5px;font-size:.8rem}"
    h = h & ".leg-dot{width:12px;height:12px;border-radius:3px}"
    h = h & ".gantt-wrap{padding:20px 30px;overflow-x:auto}"
    h = h & ".gantt-container{min-width:1200px}"
    h = h & ".gantt-header{display:grid;background:#1a2e42;border-radius:8px 8px 0 0;border:1px solid #2e4a63}"
    h = h & ".gantt-row{display:grid;border-bottom:1px solid #1e3550;transition:background .15s}"
    h = h & ".gantt-row:hover{background:rgba(46,117,182,.15)}"
    h = h & ".gantt-row.phase-row{background:linear-gradient(90deg,#1f4e79,#1a3f63);font-weight:700;font-size:.9rem}"
    h = h & ".gantt-row.subtask-row .task-info{padding-left:28px;font-style:italic;font-size:.85rem}"
    h = h & ".task-info{padding:8px 12px;border-right:1px solid #2e4a63;font-size:.88rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
    h = h & ".task-chart{position:relative;height:36px;display:flex;align-items:center}"
    h = h & ".gantt-bar{position:absolute;height:22px;border-radius:4px;display:flex;align-items:center;padding:0 6px;font-size:.75rem;font-weight:600;color:#fff;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;box-shadow:0 2px 6px rgba(0,0,0,.3);cursor:pointer;transition:filter .2s}"
    h = h & ".gantt-bar:hover{filter:brightness(1.2)}"
    h = h & ".gantt-bar .fill{position:absolute;left:0;top:0;height:100%;background:rgba(255,255,255,.2);border-radius:4px 0 0 4px}"
    h = h & ".today-line{position:absolute;top:0;bottom:0;width:2px;background:#ff4444;z-index:10;pointer-events:none}"
    h = h & ".today-line::after{content:'TODAY';position:absolute;top:-18px;left:50%;transform:translateX(-50%);font-size:.65rem;color:#ff4444;font-weight:700;white-space:nowrap}"
    h = h & ".date-header{display:flex;border-bottom:1px solid #2e4a63;background:#152030;min-height:32px}"
    h = h & ".date-col{border-right:1px solid #2e4a63;font-size:.7rem;text-align:center;padding:4px 2px;color:#93b5d0}"
    h = h & ".status-badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:.75rem;font-weight:700}"
    h = h & ".stat-Completed{background:#00b050;color:#fff}"
    h = h & ".stat-InProgress{background:#ffb900;color:#000}"
    h = h & ".stat-NotStarted{background:#a5a5a5;color:#fff}"
    h = h & ".stat-Overdue{background:#ff0000;color:#fff}"
    h = h & ".stat-Blocked{background:#7030a0;color:#fff}"
    h = h & ".tooltip{position:fixed;background:#1a2e42;border:1px solid #2e75b6;border-radius:8px;padding:12px;font-size:.82rem;pointer-events:none;z-index:999;max-width:280px;box-shadow:0 4px 20px rgba(0,0,0,.5);display:none}"
    h = h & ".tooltip strong{color:#2e75b6;display:block;margin-bottom:6px;font-size:.9rem}"
    h = h & ".tooltip .tt-row{display:flex;justify-content:space-between;gap:12px;margin:3px 0;color:#c0d5e8}"
    h = h & ".footer{text-align:center;padding:20px;color:#3a6080;font-size:.8rem;border-top:1px solid #1e3550}"
    h = h & "</style></head><body>"
    h = h & "<header><div><h1>&#128202; " & projectName & " — Gantt Chart</h1>"
    h = h & "<div class=""meta"">Generated: " & Format(Now(), "DD MMM YYYY HH:MM") & " &nbsp;|&nbsp; Exported from Excel</div></div></header>"
    h = h & "<div class=""controls"">"
    h = h & "<label>Search: <input type=""text"" id=""searchBox"" placeholder=""Filter tasks..."" oninput=""filterTasks()""></label>"
    h = h & "<label>Status: <select id=""statusFilter"" onchange=""filterTasks()"">"
    h = h & "<option value="""">All Statuses</option>"
    h = h & "<option>Completed</option><option>In Progress</option><option>Not Started</option><option>Overdue</option><option>Blocked</option>"
    h = h & "</select></label>"
    h = h & "<label>Resource: <select id=""resourceFilter"" onchange=""filterTasks()""><option value="""">All Resources</option></select></label>"
    h = h & "<div class=""legend"">"
    h = h & "<span class=""leg""><span class=""leg-dot"" style=""background:#00b050""></span>Completed</span>"
    h = h & "<span class=""leg""><span class=""leg-dot"" style=""background:#ffb900""></span>In Progress</span>"
    h = h & "<span class=""leg""><span class=""leg-dot"" style=""background:#a5a5a5""></span>Not Started</span>"
    h = h & "<span class=""leg""><span class=""leg-dot"" style=""background:#ff0000""></span>Overdue</span>"
    h = h & "<span class=""leg""><span class=""leg-dot"" style=""background:#7030a0""></span>Blocked</span>"
    h = h & "</div></div>"
    h = h & "<div class=""gantt-wrap""><div id=""ganttContainer"" class=""gantt-container""></div></div>"
    h = h & "<div class=""tooltip"" id=""tooltip""></div>"
    h = h & "<div class=""footer"">&#128202; Interactive Gantt Chart &mdash; " & projectName & " &mdash; Generated " & Format(Now(), "DD MMM YYYY") & "</div>"
    h = h & "<script>"
    h = h & "const TASKS=" & jsonTasks & ";"
    h = h & "const STATUS_COLORS={'Completed':'#00b050','In Progress':'#ffb900','Not Started':'#a5a5a5','Overdue':'#ff0000','Blocked':'#7030a0'};"
    h = h & "function parseDate(s){if(!s)return null;const[d,m,y]=s.split('/');return new Date(y,m-1,d);}"
    h = h & "function fmtDate(d){if(!d)return'';return d.toLocaleDateString('en-GB',{day:'2-digit',month:'short',year:'numeric'});}"
    h = h & "function buildGantt(tasks){"
    h = h & "const dates=tasks.filter(t=>t.start&&t.end).map(t=>[parseDate(t.start),parseDate(t.end)]).flat().filter(Boolean);"
    h = h & "if(!dates.length)return;"
    h = h & "let minDate=new Date(Math.min(...dates));let maxDate=new Date(Math.max(...dates));"
    h = h & "minDate.setDate(minDate.getDate()-7);maxDate.setDate(maxDate.getDate()+14);"
    h = h & "const totalDays=(maxDate-minDate)/(1000*86400);"
    h = h & "const INFO_WIDTH=320;const CHART_WIDTH=Math.max(900,totalDays*24);"
    h = h & "const COL_GRID=`${INFO_WIDTH}px 80px 80px ${CHART_WIDTH}px`;"
    h = h & "const container=document.getElementById('ganttContainer');"
    h = h & "container.innerHTML='';"
    h = h & "// Date header"
    h = h & "const hdr=document.createElement('div');hdr.className='gantt-header';"
    h = h & "hdr.style.gridTemplateColumns=COL_GRID;"
    h = h & "hdr.innerHTML=`<div class='task-info' style='font-weight:700;color:#93b5d0;background:#1a2e42'>Task / Description</div><div class='task-info' style='background:#1a2e42;color:#93b5d0;text-align:center'>Start</div><div class='task-info' style='background:#1a2e42;color:#93b5d0;text-align:center'>End</div>`;"
    h = h & "const dateHdr=document.createElement('div');dateHdr.className='date-header';dateHdr.style.position='relative';"
    h = h & "const today=new Date();let cur=new Date(minDate);"
    h = h & "while(cur<=maxDate){"
    h = h & "const pct=((cur-minDate)/(maxDate-minDate))*100;"
    h = h & "const col=document.createElement('div');col.className='date-col';"
    h = h & "col.style.width=(CHART_WIDTH/totalDays)+'px';col.style.flexShrink='0';"
    h = h & "if(cur.getDate()===1||cur.getDay()===1){col.textContent=cur.toLocaleDateString('en-GB',{day:'2-digit',month:'short'});}"
    h = h & "if(cur.toDateString()===today.toDateString())col.style.background='rgba(255,68,68,.3)';"
    h = h & "dateHdr.appendChild(col);cur.setDate(cur.getDate()+1);}"
    h = h & "hdr.appendChild(dateHdr);container.appendChild(hdr);"
    h = h & "tasks.forEach(t=>{"
    h = h & "const row=document.createElement('div');row.className='gantt-row'+(t.type==='Phase'?' phase-row':t.type==='Subtask'?' subtask-row':'');"
    h = h & "row.style.gridTemplateColumns=COL_GRID;"
    h = h & "row.dataset.desc=(t.desc||'').toLowerCase();row.dataset.status=t.status||'';row.dataset.resource=t.resource||'';"
    h = h & "row.innerHTML=`<div class='task-info'>${t.type==='Phase'?'&#9660; ':''}${t.desc||''}</div>`+"
    h = h & "`<div class='task-info' style='text-align:center;font-size:.8rem'>${t.start||''}</div>`+"
    h = h & "`<div class='task-info' style='text-align:center;font-size:.8rem'>${t.end||''}</div>`;"
    h = h & "const chart=document.createElement('div');chart.className='task-chart';chart.style.position='relative';"
    h = h & "// Today line"
    h = h & "const todayPct=((today-minDate)/(maxDate-minDate))*100;"
    h = h & "if(todayPct>=0&&todayPct<=100){const tl=document.createElement('div');tl.className='today-line';tl.style.left=todayPct+'%';chart.appendChild(tl);}"
    h = h & "if(t.start&&t.end&&t.type!=='Phase'){"
    h = h & "const s=parseDate(t.start),e=parseDate(t.end);"
    h = h & "if(s&&e){const left=((s-minDate)/(maxDate-minDate))*100;const width=((e-s)/(maxDate-minDate))*100;"
    h = h & "const bar=document.createElement('div');bar.className='gantt-bar';"
    h = h & "bar.style.left=Math.max(0,left)+'%';bar.style.width=Math.max(.5,width)+'%';"
    h = h & "bar.style.background=STATUS_COLORS[t.status]||'#666';"
    h = h & "const fill=document.createElement('div');fill.className='fill';fill.style.width=(t.pct*100)+'%';bar.appendChild(fill);"
    h = h & "const lbl=document.createElement('span');lbl.style.position='relative';lbl.style.zIndex='1';"
    h = h & "lbl.textContent=Math.round(t.pct*100)+'%';bar.appendChild(lbl);"
    h = h & "bar.addEventListener('mousemove',e=>{showTooltip(e,t);});bar.addEventListener('mouseleave',hideTooltip);"
    h = h & "chart.appendChild(bar);}}"
    h = h & "row.appendChild(chart);container.appendChild(row);});"
    h = h & "}"
    h = h & "function showTooltip(e,t){"
    h = h & "const tt=document.getElementById('tooltip');"
    h = h & "tt.innerHTML=`<strong>${t.desc}</strong><div class='tt-row'><span>Status</span><span class='status-badge stat-${(t.status||'').replace(' ','')}'>${t.status}</span></div><div class='tt-row'><span>Resource</span><span>${t.resource}</span></div><div class='tt-row'><span>Start</span><span>${t.start}</span></div><div class='tt-row'><span>End</span><span>${t.end}</span></div><div class='tt-row'><span>Progress</span><span>${Math.round(t.pct*100)}%</span></div>${t.notes?'<div class=\"tt-row\"><span>Notes</span><span>'+t.notes+'</span></div>':''}`;"
    h = h & "tt.style.display='block';tt.style.left=(e.clientX+12)+'px';tt.style.top=(e.clientY-10)+'px';}"
    h = h & "function hideTooltip(){document.getElementById('tooltip').style.display='none';}"
    h = h & "function filterTasks(){"
    h = h & "const q=document.getElementById('searchBox').value.toLowerCase();"
    h = h & "const s=document.getElementById('statusFilter').value;"
    h = h & "const r=document.getElementById('resourceFilter').value;"
    h = h & "document.querySelectorAll('.gantt-row').forEach(row=>{"
    h = h & "const matchQ=!q||row.dataset.desc.includes(q);"
    h = h & "const matchS=!s||row.dataset.status===s;"
    h = h & "const matchR=!r||row.dataset.resource===r;"
    h = h & "row.style.display=(matchQ&&matchS&&matchR)||row.classList.contains('phase-row')?'':'none';});}"
    h = h & "// Populate resource filter"
    h = h & "const resources=[...new Set(TASKS.filter(t=>t.resource).map(t=>t.resource))];"
    h = h & "const rf=document.getElementById('resourceFilter');"
    h = h & "resources.forEach(r=>{const o=document.createElement('option');o.value=r;o.textContent=r;rf.appendChild(o);});"
    h = h & "buildGantt(TASKS);"
    h = h & "</script></body></html>"
    BuildGanttHTML = h
End Function

'=============================================================================
' EXPORT 2: PROJECT DASHBOARD
'=============================================================================
Sub ExportDashboard()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)

    Dim projectName As String
    projectName = ws.Range("E2").Value
    If projectName = "" Then projectName = "Project Dashboard"

    Dim tasks As Variant
    tasks = CollectTasks()

    If Not IsArray(tasks) Then
        MsgBox "No task data found.", vbExclamation
        Exit Sub
    End If

    Dim jsonTasks As String
    jsonTasks = "["
    Dim i As Long
    For i = 1 To UBound(tasks, 1)
        If i > 1 Then jsonTasks = jsonTasks & ","
        Dim pct As Double
        If IsNumeric(tasks(i, 8)) Then pct = tasks(i, 8) Else pct = 0
        jsonTasks = jsonTasks & "{"
        jsonTasks = jsonTasks & """id"":" & IIf(tasks(i, 1) = "", "0", tasks(i, 1)) & ","
        jsonTasks = jsonTasks & """type"":""" & EscapeJson(CStr(tasks(i, 2))) & ""","
        jsonTasks = jsonTasks & """phase"":""" & EscapeJson(CStr(tasks(i, 3))) & ""","
        jsonTasks = jsonTasks & """desc"":""" & EscapeJson(CStr(tasks(i, 4))) & ""","
        jsonTasks = jsonTasks & """start"":""" & tasks(i, 5) & ""","
        jsonTasks = jsonTasks & """end"":""" & tasks(i, 6) & ""","
        jsonTasks = jsonTasks & """resource"":""" & EscapeJson(CStr(tasks(i, 7))) & ""","
        jsonTasks = jsonTasks & """pct"":" & pct & ","
        jsonTasks = jsonTasks & """status"":""" & EscapeJson(CStr(tasks(i, 9))) & ""","
        jsonTasks = jsonTasks & """notes"":""" & EscapeJson(CStr(tasks(i, 10))) & """"
        jsonTasks = jsonTasks & "}"
    Next i
    jsonTasks = jsonTasks & "]"

    Dim html As String
    html = BuildDashboardHTML(projectName, jsonTasks)

    Dim savePath As String
    savePath = GetSavePath("ProjectDashboard.html")
    WriteFile savePath, html
    OpenInBrowser savePath
    MsgBox "Dashboard exported to:" & vbCrLf & savePath, vbInformation, "Export Complete"
End Sub

Function BuildDashboardHTML(projectName As String, jsonTasks As String) As String
    Dim h As String
    h = "<!DOCTYPE html><html lang=""en""><head><meta charset=""UTF-8"">"
    h = h & "<title>Dashboard - " & projectName & "</title>"
    h = h & "<script src=""https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js""></script>"
    h = h & "<style>"
    h = h & "*{box-sizing:border-box;margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif}"
    h = h & "body{background:#0f1923;color:#dce8f0;min-height:100vh}"
    h = h & "header{background:linear-gradient(135deg,#1f4e79 0%,#2e75b6 100%);padding:22px 32px;box-shadow:0 3px 15px rgba(0,0,0,.4)}"
    h = h & "header h1{font-size:1.7rem;font-weight:700}header p{opacity:.75;font-size:.9rem;margin-top:4px}"
    h = h & ".kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;padding:24px 32px}"
    h = h & ".kpi{background:#162636;border-radius:12px;padding:20px;text-align:center;border:1px solid #2e4a63;transition:transform .2s}"
    h = h & ".kpi:hover{transform:translateY(-3px)}"
    h = h & ".kpi .num{font-size:2.4rem;font-weight:800;line-height:1}"
    h = h & ".kpi .lbl{font-size:.82rem;color:#7bafc8;margin-top:6px;text-transform:uppercase;letter-spacing:.5px}"
    h = h & ".charts-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;padding:0 32px 24px}"
    h = h & "@media(max-width:900px){.charts-grid{grid-template-columns:1fr 1fr}}"
    h = h & "@media(max-width:600px){.charts-grid{grid-template-columns:1fr}}"
    h = h & ".chart-card{background:#162636;border-radius:12px;padding:20px;border:1px solid #2e4a63}"
    h = h & ".chart-card h3{font-size:1rem;margin-bottom:14px;color:#93b5d0;font-weight:600}"
    h = h & ".phase-progress{padding:0 32px 32px}"
    h = h & ".phase-card{background:#162636;border-radius:10px;padding:16px 20px;margin-bottom:10px;border:1px solid #2e4a63}"
    h = h & ".phase-name{font-weight:700;margin-bottom:8px;font-size:.92rem}"
    h = h & ".phase-bar-bg{background:#0d1b2a;border-radius:6px;height:14px;overflow:hidden}"
    h = h & ".phase-bar{height:100%;border-radius:6px;background:linear-gradient(90deg,#00b050,#00e065);transition:width .5s}"
    h = h & ".phase-meta{display:flex;justify-content:space-between;font-size:.78rem;color:#7bafc8;margin-top:4px}"
    h = h & ".task-table-wrap{padding:0 32px 32px;overflow-x:auto}"
    h = h & "table{width:100%;border-collapse:collapse;font-size:.85rem}"
    h = h & "th{background:#1f4e79;color:#fff;padding:10px 12px;text-align:left;font-weight:600}"
    h = h & "td{padding:8px 12px;border-bottom:1px solid #1e3550}"
    h = h & "tr:hover td{background:rgba(46,117,182,.12)}"
    h = h & ".badge{display:inline-block;padding:2px 10px;border-radius:10px;font-size:.75rem;font-weight:700}"
    h = h & ".bg-green{background:#00b050;color:#fff}.bg-amber{background:#ffb900;color:#000}"
    h = h & ".bg-grey{background:#a5a5a5;color:#fff}.bg-red{background:#ff0000;color:#fff}.bg-purple{background:#7030a0;color:#fff}"
    h = h & ".progress-mini{width:80px;height:8px;background:#0d1b2a;border-radius:4px;overflow:hidden;display:inline-block;vertical-align:middle;margin-left:6px}"
    h = h & ".progress-mini-fill{height:100%;background:#2e75b6;border-radius:4px}"
    h = h & ".footer{text-align:center;padding:20px;color:#3a6080;font-size:.8rem;border-top:1px solid #1e3550}"
    h = h & "</style></head><body>"
    h = h & "<header><h1>&#128202; " & projectName & " &mdash; Project Dashboard</h1>"
    h = h & "<p>Generated: " & Format(Now(), "DD MMM YYYY HH:MM") & "</p></header>"
    h = h & "<div class=""kpi-grid"">"
    h = h & "<div class=""kpi""><div class=""num"" id=""kpiTotal"">-</div><div class=""lbl"">Total Tasks</div></div>"
    h = h & "<div class=""kpi"" style=""border-color:#00b050""><div class=""num"" style=""color:#00b050"" id=""kpiComplete"">-</div><div class=""lbl"">Completed</div></div>"
    h = h & "<div class=""kpi"" style=""border-color:#ffb900""><div class=""num"" style=""color:#ffb900"" id=""kpiInProg"">-</div><div class=""lbl"">In Progress</div></div>"
    h = h & "<div class=""kpi"" style=""border-color:#ff0000""><div class=""num"" style=""color:#ff0000"" id=""kpiOverdue"">-</div><div class=""lbl"">Overdue</div></div>"
    h = h & "<div class=""kpi"" style=""border-color:#7030a0""><div class=""num"" style=""color:#7030a0"" id=""kpiBlocked"">-</div><div class=""lbl"">Blocked</div></div>"
    h = h & "<div class=""kpi"" style=""border-color:#2e75b6""><div class=""num"" style=""color:#2e75b6"" id=""kpiPct"">-</div><div class=""lbl"">% Complete</div></div>"
    h = h & "</div>"
    h = h & "<div class=""charts-grid"">"
    h = h & "<div class=""chart-card""><h3>Status Breakdown</h3><canvas id=""statusChart"" height=""200""></canvas></div>"
    h = h & "<div class=""chart-card""><h3>Resource Workload</h3><canvas id=""resourceChart"" height=""200""></canvas></div>"
    h = h & "<div class=""chart-card""><h3>Phase Completion</h3><canvas id=""phaseChart"" height=""200""></canvas></div>"
    h = h & "</div>"
    h = h & "<div class=""phase-progress""><h3 style=""color:#93b5d0;margin-bottom:12px;font-size:1rem"">Phase Progress Breakdown</h3><div id=""phaseCards""></div></div>"
    h = h & "<div class=""task-table-wrap""><h3 style=""color:#93b5d0;margin-bottom:12px;font-size:1rem"">All Tasks</h3>"
    h = h & "<table><thead><tr><th>ID</th><th>Type</th><th>Description</th><th>Resource</th><th>Start</th><th>End</th><th>Progress</th><th>Status</th><th>Notes</th></tr></thead>"
    h = h & "<tbody id=""taskTableBody""></tbody></table></div>"
    h = h & "<div class=""footer"">&#128202; Project Dashboard &mdash; " & projectName & " &mdash; " & Format(Now(), "DD MMM YYYY") & "</div>"
    h = h & "<script>"
    h = h & "const TASKS=" & jsonTasks & ";"
    h = h & "const STATUS_COLORS={'Completed':'#00b050','In Progress':'#ffb900','Not Started':'#a5a5a5','Overdue':'#ff0000','Blocked':'#7030a0'};"
    h = h & "const workerTasks=TASKS.filter(t=>t.type==='Task'||t.type==='Subtask');"
    h = h & "const total=workerTasks.length;"
    h = h & "const byStatus={};"
    h = h & "workerTasks.forEach(t=>{byStatus[t.status]=(byStatus[t.status]||0)+1;});"
    h = h & "const completed=byStatus['Completed']||0;"
    h = h & "document.getElementById('kpiTotal').textContent=total;"
    h = h & "document.getElementById('kpiComplete').textContent=completed;"
    h = h & "document.getElementById('kpiInProg').textContent=byStatus['In Progress']||0;"
    h = h & "document.getElementById('kpiOverdue').textContent=byStatus['Overdue']||0;"
    h = h & "document.getElementById('kpiBlocked').textContent=byStatus['Blocked']||0;"
    h = h & "document.getElementById('kpiPct').textContent=total>0?Math.round(completed/total*100)+'%':'0%';"
    h = h & "// Status chart"
    h = h & "new Chart(document.getElementById('statusChart'),{type:'doughnut',data:{labels:Object.keys(byStatus),datasets:[{data:Object.values(byStatus),backgroundColor:Object.keys(byStatus).map(s=>STATUS_COLORS[s]||'#666'),borderWidth:0}]},options:{plugins:{legend:{labels:{color:'#dce8f0'}}},cutout:'65%'}});"
    h = h & "// Resource chart"
    h = h & "const byResource={};"
    h = h & "workerTasks.forEach(t=>{if(t.resource)byResource[t.resource]=(byResource[t.resource]||0)+1;});"
    h = h & "const sortedRes=Object.entries(byResource).sort((a,b)=>b[1]-a[1]).slice(0,10);"
    h = h & "new Chart(document.getElementById('resourceChart'),{type:'bar',data:{labels:sortedRes.map(r=>r[0]),datasets:[{label:'Tasks',data:sortedRes.map(r=>r[1]),backgroundColor:'#2e75b6',borderRadius:4}]},options:{indexAxis:'y',plugins:{legend:{display:false}},scales:{x:{ticks:{color:'#7bafc8'},grid:{color:'#1e3550'}},y:{ticks:{color:'#dce8f0'},grid:{color:'transparent'}}}}});"
    h = h & "// Phase chart & cards"
    h = h & "const phases=[...new Set(TASKS.filter(t=>t.type==='Phase').map(t=>t.phase))];"
    h = h & "const phaseComps=phases.map(p=>{const pt=TASKS.filter(t=>t.phase===p&&(t.type==='Task'||t.type==='Subtask'));return pt.length>0?Math.round(pt.filter(t=>t.status==='Completed').length/pt.length*100):0;});"
    h = h & "new Chart(document.getElementById('phaseChart'),{type:'bar',data:{labels:phases.map(p=>p.replace(/Phase \\d+: /,'')),datasets:[{label:'% Done',data:phaseComps,backgroundColor:phaseComps.map(v=>v===100?'#00b050':v>50?'#ffb900':'#2e75b6'),borderRadius:4}]},options:{plugins:{legend:{display:false}},scales:{y:{max:100,ticks:{callback:v=>v+'%',color:'#7bafc8'},grid:{color:'#1e3550'}},x:{ticks:{color:'#dce8f0',maxRotation:45},grid:{color:'transparent'}}}}});"
    h = h & "const phaseCards=document.getElementById('phaseCards');"
    h = h & "phases.forEach((p,i)=>{"
    h = h & "const pTasks=TASKS.filter(t=>t.phase===p&&(t.type==='Task'||t.type==='Subtask'));"
    h = h & "const pComp=pTasks.filter(t=>t.status==='Completed').length;"
    h = h & "const pPct=pTasks.length>0?Math.round(pComp/pTasks.length*100):0;"
    h = h & "const card=document.createElement('div');card.className='phase-card';"
    h = h & "card.innerHTML=`<div class='phase-name'>${p}</div><div class='phase-bar-bg'><div class='phase-bar' style='width:${pPct}%;background:${pPct===100?'linear-gradient(90deg,#00b050,#00e065)':pPct>50?'linear-gradient(90deg,#ffb900,#ffd000)':'linear-gradient(90deg,#2e75b6,#4a9fd6)'}'></div></div><div class='phase-meta'><span>${pComp} of ${pTasks.length} tasks complete</span><span>${pPct}%</span></div>`;"
    h = h & "phaseCards.appendChild(card);});"
    h = h & "// Task table"
    h = h & "const tbody=document.getElementById('taskTableBody');"
    h = h & "const statusBadge={'Completed':'bg-green','In Progress':'bg-amber','Not Started':'bg-grey','Overdue':'bg-red','Blocked':'bg-purple'};"
    h = h & "workerTasks.forEach(t=>{"
    h = h & "const tr=document.createElement('tr');"
    h = h & "const pct=Math.round((t.pct||0)*100);"
    h = h & "tr.innerHTML=`<td>${t.id||''}</td><td>${t.type}</td><td>${t.desc}</td><td>${t.resource}</td><td>${t.start}</td><td>${t.end}</td>`+"
    h = h & "`<td><span>${pct}%</span><span class='progress-mini'><span class='progress-mini-fill' style='width:${pct}%'></span></span></td>`+"
    h = h & "`<td><span class='badge ${statusBadge[t.status]||""}'>${t.status}</span></td><td style='color:#7bafc8'>${t.notes||''}</td>`;"
    h = h & "tbody.appendChild(tr);});"
    h = h & "</script></body></html>"
    BuildDashboardHTML = h
End Function

'=============================================================================
' EXPORT 3: RAG STATUS REPORT
'=============================================================================
Sub ExportRAGReport()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(PLAN_SHEET)

    Dim projectName As String
    projectName = ws.Range("E2").Value
    If projectName = "" Then projectName = "Project RAG Report"

    Dim tasks As Variant
    tasks = CollectTasks()

    If Not IsArray(tasks) Then
        MsgBox "No task data found.", vbExclamation
        Exit Sub
    End If

    Dim jsonTasks As String
    jsonTasks = "["
    Dim i As Long
    For i = 1 To UBound(tasks, 1)
        If i > 1 Then jsonTasks = jsonTasks & ","
        Dim pct As Double
        If IsNumeric(tasks(i, 8)) Then pct = tasks(i, 8) Else pct = 0
        jsonTasks = jsonTasks & "{"
        jsonTasks = jsonTasks & """id"":" & IIf(tasks(i, 1) = "", "0", tasks(i, 1)) & ","
        jsonTasks = jsonTasks & """type"":""" & EscapeJson(CStr(tasks(i, 2))) & ""","
        jsonTasks = jsonTasks & """phase"":""" & EscapeJson(CStr(tasks(i, 3))) & ""","
        jsonTasks = jsonTasks & """desc"":""" & EscapeJson(CStr(tasks(i, 4))) & ""","
        jsonTasks = jsonTasks & """start"":""" & tasks(i, 5) & ""","
        jsonTasks = jsonTasks & """end"":""" & tasks(i, 6) & ""","
        jsonTasks = jsonTasks & """resource"":""" & EscapeJson(CStr(tasks(i, 7))) & ""","
        jsonTasks = jsonTasks & """pct"":" & pct & ","
        jsonTasks = jsonTasks & """status"":""" & EscapeJson(CStr(tasks(i, 9))) & ""","
        jsonTasks = jsonTasks & """notes"":""" & EscapeJson(CStr(tasks(i, 10))) & """"
        jsonTasks = jsonTasks & "}"
    Next i
    jsonTasks = jsonTasks & "]"

    Dim html As String
    html = BuildRAGHTML(projectName, jsonTasks)

    Dim savePath As String
    savePath = GetSavePath("ProjectRAGReport.html")
    WriteFile savePath, html
    OpenInBrowser savePath
    MsgBox "RAG Status Report exported to:" & vbCrLf & savePath, vbInformation, "Export Complete"
End Sub

Function BuildRAGHTML(projectName As String, jsonTasks As String) As String
    Dim h As String
    h = "<!DOCTYPE html><html lang=""en""><head><meta charset=""UTF-8"">"
    h = h & "<title>RAG Report - " & projectName & "</title>"
    h = h & "<style>"
    h = h & "*{box-sizing:border-box;margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif}"
    h = h & "body{background:#fff;color:#222;min-height:100vh}"
    h = h & "@media print{.no-print{display:none}body{background:#fff}}"
    h = h & "header{background:#1f4e79;color:#fff;padding:24px 40px;print-color-adjust:exact}"
    h = h & "header h1{font-size:1.8rem;font-weight:700}"
    h = h & "header .sub{display:flex;gap:24px;margin-top:8px;font-size:.9rem;opacity:.85}"
    h = h & ".rag-summary{display:flex;gap:0;margin:20px 40px;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.15)}"
    h = h & ".rag-band{flex:1;padding:20px;text-align:center;color:#fff;print-color-adjust:exact}"
    h = h & ".rag-band .num{font-size:2.5rem;font-weight:800;line-height:1}"
    h = h & ".rag-band .lbl{font-size:.85rem;text-transform:uppercase;letter-spacing:.5px;margin-top:4px}"
    h = h & ".rag-R{background:#d00000}.rag-A{background:#e89b00}.rag-G{background:#00863f}"
    h = h & ".section-title{font-size:1.1rem;font-weight:700;margin:24px 40px 10px;color:#1f4e79;border-bottom:2px solid #1f4e79;padding-bottom:6px}"
    h = h & ".phase-section{margin:0 40px 20px}"
    h = h & ".phase-hdr{display:flex;align-items:center;gap:12px;padding:10px 16px;border-radius:8px 8px 0 0;font-weight:700;font-size:.95rem;cursor:pointer;print-color-adjust:exact}"
    h = h & ".phase-hdr.rag-R{background:#d00000;color:#fff}"
    h = h & ".phase-hdr.rag-A{background:#e89b00;color:#fff}"
    h = h & ".phase-hdr.rag-G{background:#00863f;color:#fff}"
    h = h & ".phase-hdr.rag-B{background:#1f4e79;color:#fff}"
    h = h & ".rag-dot{width:16px;height:16px;border-radius:50%;border:2px solid rgba(255,255,255,.5)}"
    h = h & ".phase-body{border:1px solid #ddd;border-top:none;border-radius:0 0 8px 8px;overflow:hidden}"
    h = h & "table{width:100%;border-collapse:collapse;font-size:.85rem}"
    h = h & "th{background:#f0f5fb;color:#1f4e79;padding:8px 12px;text-align:left;font-weight:600;border-bottom:2px solid #1f4e79}"
    h = h & "td{padding:7px 12px;border-bottom:1px solid #eee}"
    h = h & "tr:last-child td{border-bottom:none}"
    h = h & "tr.subtask td:nth-child(3){padding-left:24px;font-style:italic;color:#555}"
    h = h & ".rag-pill{display:inline-block;padding:3px 12px;border-radius:12px;font-size:.75rem;font-weight:700;color:#fff;print-color-adjust:exact}"
    h = h & ".pill-R{background:#d00000}.pill-A{background:#e89b00;color:#000}.pill-G{background:#00863f}.pill-B{background:#1f4e79}"
    h = h & ".pbar{width:80px;height:8px;background:#eee;border-radius:4px;display:inline-block;vertical-align:middle;margin-left:6px;overflow:hidden}"
    h = h & ".pbar-fill{height:100%;border-radius:4px;print-color-adjust:exact}"
    h = h & ".pbar-fill.R{background:#d00000}.pbar-fill.A{background:#e89b00}.pbar-fill.G{background:#00863f}"
    h = h & ".notes-cell{color:#e65c00;font-style:italic;font-size:.82rem}"
    h = h & ".btn-print{position:fixed;bottom:24px;right:24px;background:#1f4e79;color:#fff;border:none;padding:12px 22px;border-radius:8px;cursor:pointer;font-size:.9rem;font-weight:600;box-shadow:0 4px 12px rgba(0,0,0,.25)}"
    h = h & ".btn-print:hover{background:#2e75b6}"
    h = h & ".footer{text-align:center;padding:20px;color:#888;font-size:.8rem;border-top:1px solid #eee;margin-top:20px}"
    h = h & ".overall-rag{display:inline-flex;align-items:center;gap:10px;padding:10px 20px;border-radius:8px;font-weight:700;font-size:1.1rem;margin:16px 40px;color:#fff;print-color-adjust:exact}"
    h = h & "</style></head><body>"
    h = h & "<header>"
    h = h & "<h1>&#128994;&#128993;&#128308; RAG Status Report &mdash; " & projectName & "</h1>"
    h = h & "<div class=""sub""><span>&#128197; Generated: " & Format(Now(), "DD MMM YYYY HH:MM") & "</span>"
    h = h & "<span>&#128203; This report reflects live data from the project plan workbook.</span></div>"
    h = h & "</header>"
    h = h & "<div id=""overallRAG""></div>"
    h = h & "<div class=""rag-summary""><div class=""rag-band rag-R""><div class=""num"" id=""cntR"">-</div><div class=""lbl"">&#128308; Red — Overdue / Blocked</div></div>"
    h = h & "<div class=""rag-band rag-A""><div class=""num"" id=""cntA"">-</div><div class=""lbl"">&#128993; Amber — In Progress</div></div>"
    h = h & "<div class=""rag-band rag-G""><div class=""num"" id=""cntG"">-</div><div class=""lbl"">&#128994; Green — Completed</div></div></div>"
    h = h & "<div id=""phaseSections""></div>"
    h = h & "<button class=""btn-print no-print"" onclick=""window.print()"">&#128424; Print / Save PDF</button>"
    h = h & "<div class=""footer"">RAG Status Report &mdash; " & projectName & " &mdash; Confidential &mdash; " & Format(Now(), "DD MMM YYYY") & "</div>"
    h = h & "<script>"
    h = h & "const TASKS=" & jsonTasks & ";"
    h = h & "function getRAG(status){"
    h = h & "if(status==='Overdue'||status==='Blocked')return'R';"
    h = h & "if(status==='In Progress'||status==='Not Started')return'A';"
    h = h & "if(status==='Completed')return'G';"
    h = h & "return'B';}"
    h = h & "const phases=[...new Set(TASKS.filter(t=>t.type==='Phase').map(t=>t.phase))];"
    h = h & "const workerTasks=TASKS.filter(t=>t.type==='Task'||t.type==='Subtask');"
    h = h & "let cntR=0,cntA=0,cntG=0;"
    h = h & "workerTasks.forEach(t=>{const r=getRAG(t.status);if(r==='R')cntR++;else if(r==='A')cntA++;else if(r==='G')cntG++;});"
    h = h & "document.getElementById('cntR').textContent=cntR;"
    h = h & "document.getElementById('cntA').textContent=cntA;"
    h = h & "document.getElementById('cntG').textContent=cntG;"
    h = h & "const total=workerTasks.length;"
    h = h & "const overallRag=cntR>0?'R':cntA>0?'A':'G';"
    h = h & "const ragLabel={'R':'RED — Project has critical issues requiring immediate attention','A':'AMBER — Project in progress, some items need monitoring','G':'GREEN — Project on track, all tasks completed or progressing well'};"
    h = h & "const ragBg={'R':'#d00000','A':'#e89b00','G':'#00863f'};"
    h = h & "document.getElementById('overallRAG').innerHTML=`<div class='overall-rag' style='background:${ragBg[overallRag]}'><span style='font-size:1.4rem'>${overallRag==='R'?'&#128308;':overallRag==='A'?'&#128993;':'&#128994;'}</span> OVERALL STATUS: ${ragLabel[overallRag]}</div>`;"
    h = h & "const container=document.getElementById('phaseSections');"
    h = h & "container.innerHTML='<div class=""section-title"">Phase-by-Phase RAG Breakdown</div>';"
    h = h & "phases.forEach(phase=>{"
    h = h & "const pTasks=TASKS.filter(t=>t.phase===phase&&(t.type==='Task'||t.type==='Subtask'));"
    h = h & "const pRags=pTasks.map(t=>getRAG(t.status));"
    h = h & "const phaseRag=pRags.includes('R')?'R':pRags.includes('A')?'A':pRags.length>0?'G':'B';"
    h = h & "const ragIcon={'R':'&#128308;','A':'&#128993;','G':'&#128994;','B':'&#9898;'};"
    h = h & "const phaseSection=document.createElement('div');phaseSection.className='phase-section';"
    h = h & "let tableRows='';"
    h = h & "pTasks.forEach((t,i)=>{"
    h = h & "const tRag=getRAG(t.status);"
    h = h & "const pct=Math.round((t.pct||0)*100);"
    h = h & "const barColor=tRag==='R'?'R':tRag==='A'?'A':'G';"
    h = h & "tableRows+=`<tr class='${t.type===\"Subtask\"?\"subtask\":\"\"}'>`+"
    h = h & "`<td>${t.id||''}</td><td>${t.type}</td><td>${t.desc}</td><td>${t.resource}</td>`+"
    h = h & "`<td>${t.start}</td><td>${t.end}</td>`+"
    h = h & "`<td>${pct}%<span class='pbar'><span class='pbar-fill ${barColor}' style='width:${pct}%'></span></span></td>`+"
    h = h & "`<td><span class='rag-pill pill-${tRag}'>${tRag==='R'?'&#128308; RED':tRag==='A'?'&#128993; AMBER':'&#128994; GREEN'}</span></td>`+"
    h = h & "`<td class='notes-cell'>${t.notes||''}</td></tr>`;"
    h = h & "});"
    h = h & "phaseSection.innerHTML=`<div class='phase-hdr rag-${phaseRag}' onclick='this.nextElementSibling.style.display=this.nextElementSibling.style.display===\"none\"?\"\":\"none\"'>`+"
    h = h & "`<span>${ragIcon[phaseRag]}</span><span>${phase}</span>`+"
    h = h & "`<span style='margin-left:auto;font-size:.8rem;opacity:.8'>${pTasks.length} tasks &mdash; click to collapse</span></div>`+"
    h = h & "`<div class='phase-body'><table><thead><tr><th>ID</th><th>Type</th><th>Description</th><th>Resource</th><th>Start</th><th>End</th><th>Progress</th><th>RAG</th><th>Notes / Blockers</th></tr></thead>`+"
    h = h & "`<tbody>${tableRows}</tbody></table></div>`;"
    h = h & "container.appendChild(phaseSection);});"
    h = h & "</script></body></html>"
    BuildRAGHTML = h
End Function
'''


# ─── README ──────────────────────────────────────────────────────────────────
README_CONTENT = """# Project Management Tool — User Guide
## Version 1.0 | Excel Macro-Enabled Workbook

---

## Overview

This workbook is a comprehensive project management tool built for Excel.
It provides:

- **12 collapseable project phases** with Tasks and Subtasks
- **Auto-updating Task IDs** via VBA macro
- **UK Holiday-aware scheduling** (2025, 2026, 2027)
- **Resource management** with dropdown assignment
- **Status tracking** with colour-coded RAG indicators
- **Three one-click HTML report exports**: Gantt Chart, Project Dashboard, RAG Status Report

---

## Files Included

| File | Purpose |
|------|---------|
| `ProjectManagementTool.xlsx` | Main workbook — open this in Excel |
| `ProjectManagementTool_VBA.bas` | VBA macro code — import once into Excel |
| `README.md` | This guide |

---

## Setup Instructions (One-Time)

### Step 1 — Open the Workbook
Open `ProjectManagementTool.xlsx` in Microsoft Excel (2016 or later recommended).

### Step 2 — Save as Macro-Enabled (.xlsm)
The workbook must be saved as `.xlsm` to support VBA macros:
1. Go to **File → Save As**
2. In the "Save as type" dropdown, select **Excel Macro-Enabled Workbook (*.xlsm)**
3. Click **Save**

### Step 3 — Import the VBA Module
1. Press **Alt + F11** to open the Visual Basic Editor
2. In the menu, go to **File → Import File...**
3. Browse to and select `ProjectManagementTool_VBA.bas`
4. Click **Open** — the module will appear in the Project pane
5. Close the VBA Editor (Alt + F4 or close the window)

### Step 4 — Enable Macros
When prompted by Excel, click **Enable Content** or **Enable Macros**.
The tool will not function without macros enabled.

### Step 5 — Run Initial Setup
1. Go to **Developer → Macros** (or press Alt + F8)
2. Select **SetupProjectUI** and click **Run**
3. This applies conditional formatting and validates the sheet structure

---

## Sheet Reference

### Project Plan Sheet
The main working sheet. Contains:

| Column | Name | Description |
|--------|------|-------------|
| A | Task ID | Auto-assigned sequential number. Never edit manually. |
| B | Task Type | Dropdown: **Task**, **Subtask**, or **Phase** (section header) |
| C | Phase | The phase this task belongs to (auto-populated from section header) |
| D | Description | Task or subtask name/description |
| E | Planned Start | Start date in DD/MM/YYYY format |
| F | Planned End | End date in DD/MM/YYYY format |
| G | Resource | Assigned team member — dropdown from Resources sheet |
| H | % Done | Progress as decimal: 0 = 0%, 0.5 = 50%, 1 = 100% |
| I | Status | Dropdown: Not Started / In Progress / Completed / Overdue / Blocked |
| J | Notes / Blockers | Free text for notes, issues, dependencies |

**Project Summary (rows 1–9):**
- Rows 1–7: Auto-calculated KPIs (total tasks, completed, in progress, overdue, blocked, % complete)
- Row 9: Export buttons — click to generate HTML reports

### Resources Sheet
Add team members here BEFORE assigning them to tasks.

| Column | Description |
|--------|-------------|
| Name | Full name — must match exactly when assigning in Project Plan |
| Department | Team or department name |
| Role | Job title or role |
| Email / Contact | Contact information |

### Holidays Sheet
Pre-populated with UK public holidays for 2025, 2026, and 2027.

| Column | Description |
|--------|-------------|
| Holiday Name | Name of the public holiday or leave |
| Start Date | First date of the holiday (DD/MM/YYYY) |
| End Date | Last date of the holiday (same as start for single-day events) |
| Resource | Leave blank for public holidays (affects everyone). Enter a name for individual leave. |

> **Adding Individual Leave:** Enter the resource's name in column D.
> The scheduling macro will skip that date only for that person.

### Stakeholders Sheet
Used for email distribution when exporting reports.

| Column | Description |
|--------|-------------|
| Name | Stakeholder display name |
| Email | Full email address |
| Type | To, CC, or BCC |
| Organisation | Company or team |

---

## Working with Phases (Collapseable Sections)

Each **Phase** row is a section header. The tasks beneath it are grouped and can be collapsed:

- Click the **minus (–)** button on the left row numbers to collapse a phase
- Click the **plus (+)** button to expand it
- Use the **1** and **2** buttons at the top-left to collapse/expand all groups

**To add a new Phase:**
1. Insert a row at the top of where you want the new phase
2. Set **Task Type = Phase** in column B
3. Enter the phase name in column D
4. The VBA macro will automatically group the rows beneath it

---

## Task ID Management

Task IDs are automatically managed by the **RenumberIDs** VBA macro.

**Inserting a task:**
1. Right-click the row where you want to insert above/below
2. Select **Insert** from the context menu
3. Run **RenumberIDs** macro (Alt + F8 → RenumberIDs → Run)
4. All IDs will renumber sequentially

**Deleting a task:**
1. Select the row and press the Delete key (or right-click → Delete)
2. Run **RenumberIDs** to close the gap

> **Never manually edit column A** — it is managed entirely by the VBA macro.

---

## Status & Colour Coding

| Status | Colour | Meaning |
|--------|--------|---------|
| Not Started | Grey | Task has not begun |
| In Progress | Amber/Gold | Work is actively underway |
| Completed | Green | Task is finished |
| Overdue | Red | Past end date and not completed |
| Blocked | Purple | Cannot proceed — dependency or issue |

The % Done column should be entered as a decimal:
- `0` = 0%
- `0.25` = 25%
- `0.5` = 50%
- `0.75` = 75%
- `1` = 100%

---

## Holiday-Aware Date Scheduling

The workbook includes UK public holidays for 2025–2027 in the Holidays sheet.

**Auto-adjust dates:**
1. Press **Alt + F8**
2. Select **AdjustDatesForHolidays**
3. Click **Run**

This scans every task's start and end date and moves any that fall on a weekend or
UK public holiday to the next available working day.

**Adding custom holidays or leave:**
- Add a row to the Holidays sheet with the Name, Start Date, End Date
- Leave the Resource column blank for company-wide holidays
- Enter a resource name for individual leave (e.g. "David Patel")

---

## Exporting HTML Reports

The three export buttons on row 9 of the Project Plan sheet generate self-contained
HTML files. These open automatically in your default browser and can be shared
with stakeholders who do not have Excel.

### How to Assign Macros to Buttons

The export button labels are styled cells in row 9. To make them clickable:

1. Go to **Insert → Shapes → Rounded Rectangle**
2. Draw a shape over the button label area
3. Right-click the shape → **Format Shape**
   - Fill: Solid colour `#1F4E79` (navy)
   - Line: `#2E75B6` (blue)
4. Right-click → **Edit Text**, type the button label
5. Set font to Segoe UI, Bold, 11pt, White
6. Right-click → **Assign Macro**
7. Select the macro from the list and click OK

| Button | Macro to Assign |
|--------|----------------|
| 📊 EXPORT GANTT CHART | `ExportGanttChart` |
| 📋 EXPORT DASHBOARD | `ExportDashboard` |
| 🚦 EXPORT RAG REPORT | `ExportRAGReport` |

Alternatively, run any macro directly via **Alt + F8 → Select Macro → Run**.

---

## Macro Reference

| Macro | Purpose |
|-------|---------|
| `SetupProjectUI` | Applies conditional formatting and dropdowns. Run once after setup. |
| `RenumberIDs` | Renumbers all Task IDs sequentially. Run after inserting/deleting rows. |
| `AdjustDatesForHolidays` | Moves dates falling on weekends/holidays to next working day. |
| `ExportGanttChart` | Generates `ProjectGanttChart.html` — interactive Gantt with search & filters |
| `ExportDashboard` | Generates `ProjectDashboard.html` — KPI cards, charts, resource workload |
| `ExportRAGReport` | Generates `ProjectRAGReport.html` — phase-by-phase RAG status, printable |

---

## HTML Report Features

### Gantt Chart (`ProjectGanttChart.html`)
- Interactive timeline with coloured task bars
- Search by task name
- Filter by Status or Resource
- % Complete shown as fill within each bar
- Hover tooltip with full task details
- Today marker (red vertical line)
- Phase group headers always visible

### Project Dashboard (`ProjectDashboard.html`)
- KPI summary cards (total, completed, in progress, overdue, blocked, % complete)
- Status breakdown donut chart (Chart.js)
- Resource workload bar chart (top 10 resources)
- Phase completion bar chart
- Phase-by-phase progress bars
- Full task table with status badges and progress bars

### RAG Status Report (`ProjectRAGReport.html`)
- Overall project RAG (Red / Amber / Green) prominently displayed
- RAG count summary banner
- Phase-by-phase breakdown with collapseable sections
- Each task shows RAG rating, progress bar, and notes/blockers
- Print-friendly layout (File → Print or Ctrl+P)
- "Print / Save PDF" button in bottom-right corner

---

## Tips & Best Practices

1. **Always set Task Type first** when adding a new row — set Phase/Task/Subtask before filling other columns
2. **Add resources to the Resources sheet** before assigning them to tasks
3. **Save before exporting** — reports save to the same folder as the workbook
4. **Run AdjustDatesForHolidays** after importing this workbook or changing holidays
5. **Run RenumberIDs** after inserting or deleting task rows
6. **Do not edit column A (ID)** — it is managed by VBA
7. **Enter % Done as decimals** (0 to 1), not percentages (0% to 100%)
8. **Use Notes/Blockers column** to log issues, dependencies, and risks

---

## Troubleshooting

| Issue | Solution |
|-------|----------|
| Macros not working | Re-enable macros via File → Options → Trust Center → Trust Center Settings → Macro Settings → Enable all macros |
| Resource dropdown empty | Add resources to the Resources sheet first |
| HTML report not opening | Check your browser is set as default; or open the .html file manually from the save folder |
| IDs are out of sequence | Run the `RenumberIDs` macro (Alt + F8 → RenumberIDs → Run) |
| Dates fall on weekends/holidays | Run `AdjustDatesForHolidays` macro |
| Report saves to wrong location | Save the workbook first — reports save to the same directory as the .xlsx/.xlsm file |
| Collapsed sections won't expand | Click the + button on the left row margin, or use the 1/2 outline level buttons in the top-left |

---

## Support & Customisation

**To add more phases:** Insert rows, set Task Type = Phase for the header row, then group the task rows beneath (Data → Group → Rows).

**To add more resources:** Add rows to the Resources sheet. The Resource dropdown in Project Plan will update automatically.

**To add holidays:** Add rows to the Holidays sheet. Run AdjustDatesForHolidays to re-check all dates.

**To customise report colours:** Edit the VBA code in the `BuildGanttHTML`, `BuildDashboardHTML`, and `BuildRAGHTML` functions. Colours are CSS hex values embedded in the HTML strings.

---

*Project Management Tool — Version 1.0*
*Designed for Microsoft Excel 2016 and later*
"""


# ─── MAIN GENERATOR ──────────────────────────────────────────────────────────
def main():
    print("Building task list...")
    tasks, project_start = build_task_list()
    print(f"  → {len(tasks)} tasks across {len(PHASES_DATA)} phases")

    print("Creating workbook...")
    wb = Workbook()

    # Create all sheets
    wb.active.title = "Project Plan"
    wb.create_sheet("Resources")
    wb.create_sheet("Holidays")
    wb.create_sheet("Stakeholders")

    print("Building Project Plan sheet...")
    last_row = setup_project_plan(wb, tasks, project_start)
    print(f"  → Data rows: 12 to {last_row}")

    print("Building Resources sheet...")
    setup_resources(wb)

    print("Building Holidays sheet...")
    setup_holidays(wb)

    print("Building Stakeholders sheet...")
    setup_stakeholders(wb)

    # Save workbook
    output_path = "ProjectManagementTool.xlsx"
    print(f"Saving workbook to {output_path}...")
    wb.save(output_path)
    print(f"  → Saved: {output_path}")

    # Write VBA code
    vba_path = "ProjectManagementTool_VBA.bas"
    print(f"Writing VBA module to {vba_path}...")
    with open(vba_path, "w", encoding="utf-8") as f:
        f.write(VBA_CODE)
    print(f"  → Saved: {vba_path}")

    # Write README
    readme_path = "ProjectManagementTool_README.md"
    print(f"Writing README to {readme_path}...")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(README_CONTENT)
    print(f"  → Saved: {readme_path}")

    print("\n✅ All files generated successfully!")
    print(f"\n  📊  {output_path}")
    print(f"  📝  {vba_path}")
    print(f"  📋  {readme_path}")
    print("\nNext steps:")
    print("  1. Open ProjectManagementTool.xlsx in Excel")
    print("  2. Save As → Excel Macro-Enabled Workbook (.xlsm)")
    print("  3. Import ProjectManagementTool_VBA.bas via Alt+F11 → File → Import File")
    print("  4. Enable macros when prompted")
    print("  5. Run SetupProjectUI macro (Alt+F8) once to apply conditional formatting")


if __name__ == "__main__":
    main()
